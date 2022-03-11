#----------------------------------------------------------------------------------------
# Coded by SONG Sungyoon, SHIN Heewon 2022/02
# Institute: KERI
#----------------------------------------------------------------------------------------
import os, sys
PSSE_PATH=r'C:\Program Files\PTI\PSSE35\35.2\PSSPY37'
sys.path.append(PSSE_PATH)
os.environ['PATH'] += ';' + PSSE_PATH
import psspy
from psspy import _i
from psspy import _f
from psspy import _s
import redirect
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
#----------------------------------------------------------------------------------------
# PSSE ½ÇÇà
redirect.psse2py()
psspy.psseinit(15000)
_i,_f = psspy.getbatdefaults()
output_dir = os.getcwd()
working_dir = output_dir + os.sep
#----------------------------------------------------------------------------------------
# Progress ½ÇÇà
psspy.lines_per_page_one_device(1,60)
psspy.progress_output(2,os.path.join(working_dir,'PROGRESS.txt'),[0,0])
psspy.lines_per_page_one_device(1,60)
psspy.prompt_output(2,os.path.join(working_dir,'log.txt'),[0,0])
psspy.lines_per_page_one_device(1,60)
psspy.report_output(2,os.path.join(working_dir,'REPORT.txt'),[0,0])
psspy.lines_per_page_one_device(1,60)
psspy.alert_output(2,os.path.join(working_dir,'log.txt'),[0,0])
#----------------------------------------------------------------------------------------
working_dir=r"C:\Users\yungun\Desktop\labsil\윤민한교수님+강화학습\연구" + os.sep

psspy.case(r"""C:\Users\yungun\Desktop\labsil\윤민한교수님+강화학습\연구\IEEE39.sav""")
psspy.fdns([0,0,0,0,0,0,99,0])
psspy.cong(0)
psspy.conl(0,1,1,[0,0],[ 100.0,0.0,0.0, 100.0])
psspy.conl(0,1,2,[0,0],[ 100.0,0.0,0.0, 100.0])
psspy.conl(0,1,3,[0,0],[ 100.0,0.0,0.0, 100.0])
psspy.ordr(1)
psspy.fact()
psspy.tysl(0)

psspy.dyre_new([1,1,1,1],r"""C:\Users\yungun\Desktop\labsil\윤민한교수님+강화학습\연구\IEEE39_IEESGOD.dyr""","","","")
time_step = 0.005
psspy.dynamics_solution_param_2([_i,_i,_i,_i,_i,_i,_i,_i],[_f,_f, time_step,_f,_f,_f,_f,_f])
psspy.change_channel_out_file(r"""C:\Users\yungun\Desktop\labsil\윤민한교수님+강화학습\연구\example.out""")
psspy.chsb(0,1,[-1,-1,-1,1,12,0])
psspy.strt_2([0,0],r"""C:\Users\yungun\Desktop\labsil\윤민한교수님+강화학습\연구\example.out""")

psspy.run(0, 1.0,1,1,0)
LOAD = [0.804328802, 0.802747521, ]
for i in range(1,round(300/time_step)):
    psspy.run(0, 1.0+time_step*i,1,1,0)
    if (1.0+time_step)*i % 50 == 0
    psspy.load_chng_6(3 ,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 322.0*self.LOAD[self.time] , 2.40*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(4 ,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 500.0*self.LOAD[self.time] ,184.0*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(7 ,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 233.8*self.LOAD[self.time] , 84.0*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(8 ,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 522.0*self.LOAD[self.time] ,176.0*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(12,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 7.500*self.LOAD[self.time] , 88.0*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(15,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 320.0*self.LOAD[self.time] ,153.0*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(16,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 329.0*self.LOAD[self.time] , 32.3*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(18,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 158.0*self.LOAD[self.time] , 30.0*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(20,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 628.0*self.LOAD[self.time] ,103.0*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(21,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 274.0*self.LOAD[self.time] ,115.0*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(23,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 247.5*self.LOAD[self.time] , 84.6*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(24,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 308.6*self.LOAD[self.time] ,(-92)*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(25,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 224.0*self.LOAD[self.time] , 47.2*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(26,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 139.0*self.LOAD[self.time] , 17.0*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(27,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 281.0*self.LOAD[self.time] , 75.5*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(28,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 206.0*self.LOAD[self.time] , 27.6*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(29,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 283.5*self.LOAD[self.time] , 26.9*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(31,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[ 9.200*self.LOAD[self.time] , 4.60*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    psspy.load_chng_6(39,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[1104.0*self.LOAD[self.time] ,250.0*self.LOAD[self.time],_f,_f,_f,_f,_f,_f],"")
    PVcor = self.PV[self.time]*(1+((np.random.rand()-0.5)*2)*0.2)
    WTcor = self.Wind[self.time]*(1+((np.random.rand()-0.5)*2)*0.3)

        #신재생에너지 
    psspy.load_data_6(3 ,r"""2""",[_i,_i,_i,_i,_i,_i,_i], [-500.0*PVcor,_f,_f,_f,_f,_f,_f,_f],"")
    psspy.load_data_6(8 ,r"""2""",[_i,_i,_i,_i,_i,_i,_i], [-500.0*PVcor,_f,_f,_f,_f,_f,_f,_f],"")
    psspy.load_data_6(16,r"""2""",[_i,_i,_i,_i,_i,_i,_i], [-500.0*PVcor,_f,_f,_f,_f,_f,_f,_f],"")
    psspy.load_data_6(18,r"""2""",[_i,_i,_i,_i,_i,_i,_i], [-500.0*WTcor,_f,_f,_f,_f,_f,_f,_f],"")
    psspy.load_data_6(23,r"""2""",[_i,_i,_i,_i,_i,_i,_i], [-500.0*WTcor,_f,_f,_f,_f,_f,_f,_f],"")
    psspy.load_data_6(27,r"""2""",[_i,_i,_i,_i,_i,_i,_i], [-500.0*WTcor,_f,_f,_f,_f,_f,_f,_f],"")
    if i 

    ierr, fre_senegal = psspy.chnval(34)
    fre_senegal = 60*(1-fre_senegal)
    print(fre_senegal)
#----------------------------------------------------------------------------------------
#Timestep values 
totalstep = 8500
timestep = 0.02 
simtime = 0
time= 0
readtime, readtime2 = 0

PV_freq, GEN_freq, ESS_freq=[],[],[]
P_i_target_list, P_GEN_list,  P_GEN_list_new, P_i_ESS_list = [], [], [0,0], []
#----------------------------------------------------------------------------------------
# PV data load
wb = openpyxl.load_workbook('DATA.xlsx',data_only=False)
sheet=wb['Sheet1']

#for i in range (1,int(totalstep)+1) :
#for i in range (8501,17001) :
#for i in range (17002,25502) :

for i in range (25503,34004) :
    A = sheet.cell(row=i+1, column=2).value
    PV_freq.append(A)

# Freq data load    
#for i in range (1,int(totalstep)+1) :
#for i in range (8501,17001) :
#for i in range (17002,25502) :
#for i in range (25503,34004) :   
##    B = sheet.cell(row=i, column=3).value
##    C = sheet.cell(row=i, column=4).value
##    GEN_freq.append(B)
##    ESS_freq.append(C)
    
#----------------------------------------------------------------------------------------
### AGC initialization
##f0 = 50
##B = 2.5 / 0.1
##
### Target Generator 
##gen_list = [13004,13004]
##gen_ID = [1,2]
##
### Gen info
##PMAX = [15.5,15.5]
##PMIN = [5, 5]

#----------------------------------------------------------------------------------------
# Add dummy governor without droop and time constants
##psspy.add_plant_model(13004,"1",7,r"""IEESGO""",0,"",0,[],[],11,[0.0,0.0,0.05,0.0,0.0,0.0,0.0,0.0,0.0,1.0,0.0])
##psspy.add_plant_model(13004,"2",7,r"""IEESGO""",0,"",0,[],[],11,[0.0,0.0,0.05,0.0,0.0,0.0,0.0,0.0,0.0,1.0,0.0])


# Dynamic Simulation
while simtime < totalstep-timestep :   
    psspy.run(0, simtime,0,0,0) 

    
    #PV change
    if time % 50 == 0 : # 1/0.02sec


        readtime+=1


##    #ESS
##    ESS = ESS_freq[time] - f0      
##    ESSmax= 5    
##    
##    if time % 2 == 0 :        
##        
##        if ESS <= -0.4 : #underfreq
##            
##           P_i_target_ESS =  -10*B*(ESS_freq[readtime2]-f0+0.4)
##           
##           if abs(P_i_target_ESS) > ESSmax :
##              P_i_target_ESS = ESSmax
##
##            else :
##              P_i_target_ESS = P_i_target_ESS
##              
##           psspy.change_plmod_var(11006,r"""1""",r"""CBEST""",1, P_i_target_ESS)
##           
##           # save result
##           temp = P_i_target_ESS
##           
##        if ESS >= 0.4 : #underfreq
##
##           P_i_target_ESS =  -10*B*(ESS_freq[readtime2]-f0-0.4)
##           
##           if abs(P_i_target_ESS) > ESSmax :
##              P_i_target_ESS = -ESSmax
##            
##            else :
##              P_i_target_ESS = P_i_target_ESS
##              
##           psspy.change_plmod_var(11006,r"""1""",r"""CBEST""",1, P_i_target_ESS)
##           
##           temp = P_i_target_ESS
##
##        if abs(ESS) <= 0.4 :
##
##            P_i_target_ESS =  0
##            
##            psspy.change_plmod_var(11006,r"""1""",r"""CBEST""",1, P_i_target_ESS)
##           
##            temp = 0
##
##        readtime2+=1
## 
##        P_i_ESS_list.append(temp)
##        
##    sheet.cell(row=time+2, column=12).value = temp
##
##
###----------------------------------------------------------------------------------------    
##    #AGC
##
##    # Read current P value
##    if round(simtime,4) % 90000 == 0 :
##        
##        for l in range(len(gen_list)):
##            ierr, cmpval = psspy.macdt2(gen_list[l],str(gen_ID[l]),"PQ")
##            
##            P_GEN = cmpval.real
##            
##            if round(simtime,4) % 90000 == 0 :  
##                P_GEN_list.append(P_GEN)
##
##            else :
##                P_GEN_list[l] = P_GEN_list_new[l] 
##            
##        P_GEN_total = sum(P_GEN_list)    
##
##        
##        for j in range(len(gen_list)):
##            ACE = -10*B*(GEN_freq[time]-f0)
##            
##            P_i_target = ACE * (P_GEN_list[j]/P_GEN_total)            
##            
##            if ACE >= 0 : #underfreq
##                if abs(P_i_target) >= PMAX[j]- P_GEN_list[j]:
##                    P_i_target = PMAX[j]
##
##                else :                
##                    P_i_target = P_GEN_list[j] + P_i_target
##
##
##            if ACE < 0 : #overfreq
##                if abs(P_i_target) >= P_GEN_list[j] - PMIN[j]  :
##                    P_i_target = PMIN[j]
##
##                else :                
##                    P_i_target = P_GEN_list[j] + P_i_target
##
##              
##            psspy.increment_gref(13004,"1", P_i_target/1000)
##            psspy.increment_gref(13004,"2", P_i_target/1000)
##
##            P_i_target_list.append(P_i_target)
##
##                 
##        # GEN power update
##        for g in range (len(gen_list)):            
##            P_GEN_list_new[g] = P_i_target_list[g]
##            
##            if P_GEN_list_new[g] > PMAX[g] :
##               P_GEN_list_new[g] = PMAX[g]
##
##            if P_GEN_list_new[g] <= PMIN[g] :
##               P_GEN_list_new[g] = PMIN[g]               
##                
##          
##        P_i_target_list, P_i_ESS_list  = [], []
##        
##    sheet.cell(row=time+2, column=10).value = P_GEN_list_new[0]    
##
#----------------------------------------------------------------------------------------   
    #Timestep update
    simtime+=0.02
    time+=1

##sheet.cell(row = time+1, column = 10).value= temp
##wb.save("Results.xlsx")
##wb.close

if simtime<totalstep-timestep :
    psspy.run(0, totalstep-timestep,0,0,0)


print("complete")



