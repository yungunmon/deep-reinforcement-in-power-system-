

import schedule
import time


def job():

    from selenium import webdriver
    from bs4 import BeautifulSoup
    from openpyxl import Workbook
    import openpyxl
    import re
    import datetime
    import os
    import shutil
    import timedelta
    import threading
    import time
    folder_name= 'svr'
    #======== 셀 레 니 움 설 정 ========#
    options = webdriver.ChromeOptions()
    #헤들리스로 만들기-----------------
    #options.add_argument('headless')
    #options.add_argument('--disable-gpu')
    #options.add_argument('lang=ko_KR')
    #----------------------------------
    path = 'C:/Users/svr/chromedriver.exe'
    driver = webdriver.Chrome(path,chrome_options=options)
    #======== 셀 레 니 움 설 정 ========#
    #폴더 생성을 위해 오늘의 날짜를 받아옴
    TODAY_DATE = datetime.datetime.today()
    ONEDAY_AGO = (TODAY_DATE - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
    TTODAY_DATE = TODAY_DATE.strftime("%m/%d/%Y")
    THREE_DAYS_AGO = (TODAY_DATE - datetime.timedelta(days=3)).strftime("%m/%d/%Y")
    #키워드 생성 알고리즘#
    Keyword_list1=['남동발전', '남부발전','서부발전', '중부발전','동서발전','한국전력','전력거래소','한국수력원자력'
                   ,'전력산업','발전소','석탄','복합발전','LNG','수력','SMP','원전','REC'
                   ,'신재생','태양광','풍력','연료전지','전기차','에너지저장장치' ,'우드펠릿','온실가스','미세먼지', '수소','바이오매스', '에너지'
                   ,'국제유가','환율','골드만삭스','기재부','산업부','환경부']
    report_list=['포스코경영연구원','LG경제연구원','한전경제연구원','현대경제연구원','수출입은행해외경제연구소','코트라','삼성경제연구소'
                ,'월간수소경제','한국개발연구원']
    #키워드의 엑셀파일을 만드는 알고리즘#    
    '''i = 0
    for i in range(len(Keyword_list1)):
        if os.path.isfile('C:/Users/'+folder_name+'/'+Keyword_list1[i]+'.xlsx') == False:
            wb = Workbook()
            wb.save('C:/Users/'+folder_name+'/'+Keyword_list1[i]+'.xlsx')
            TODAY_DATE = datetime.datetime.today()
    
    i = 0    
    for i in range(len(report_list)):
        if os.path.isfile('C:/Users/'+folder_name+'/'+report_list[i]+'.xlsx')==False:
            wb = Workbook()
            wb.save('C:/Users/'+folder_name+'/'+report_list[i]+'.xlsx')
            '''
    def news_crowler(Keyword_list) : 
        i = 0 ## 다음 키워드로 넘겨주는 변수
        for i in range(len(Keyword_list)): #리스트의 키워드를 반복하기 위한 For 반복문#
            #키워드의 엑셀파일을 열고, 
            wb = openpyxl.load_workbook('C:/Users/'+folder_name+'/'+Keyword_list[i]+'.xlsx')
            wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
            Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
        
            link_partition_1 = 'https://www.google.com/search?q='
            link_partition_2 = Keyword_list[i]
            link_partition_3 = ', -배구 , -손흥민, -농구, -야구, -리버풀, -핵무기, -공포, -감독&tbs=qdr:d,lr:lang_1ko&tbm=nws&source=lnt&lr=lang_ko&sa=X&ved=0ahUKEwjUuZyC1YvkAhWiyIsBHVvCC-8QpwUIHg&biw=1536&bih=754&dpr=1.25'
                 #link_partition_3은 국내 뉴스만 검색하게끔 만듦.
            Current_URL = link_partition_1 + link_partition_2 + link_partition_3
            crowling_count = 0 #크롤링을 한 횟수를 카운트하기 위해 만들어주는 변수 매번 0으로 갱신
            row = 0 #엑셀 파일을 이어붙이기위해 매번 row를 0으로 갱신
            row_index = 0
                
            while (1):
                driver.get(Current_URL)
                html = driver.page_source    
                soup = BeautifulSoup(html,'html.parser')  
                driver.implicitly_wait(300)
                #크롤링한 데이터를 엑셀에 옮겨적는 알고리즘
                for row_index in range(len(soup.select("h3 > a"))):
                    titles = soup.select("h3 > a")[row_index].text
                    links = soup.select("h3 > a")[row_index].get('href')
                    news =  soup.select("div.slp > span")[3*row_index].text
                    news = news.strip()
                    news = news.replace('(풍자)', '')
                    news = news.replace('(보도자료)', '')
                    news = news.replace('(블로그)', '')
                    Today_Sheet.cell(row=row_index+1 + crowling_count, column=1).value = titles
                    Today_Sheet.cell(row=row_index+1 + crowling_count, column=2).value = links
                    Today_Sheet.cell(row=row_index+1 + crowling_count, column=3).value = news
                    Today_Sheet.cell(row=row_index+1 + crowling_count, column=4).value = ONEDAY_AGO
                    #페이지 넘기기 알고리즘
                time.sleep(15)
                if len(soup.select('td > a.pn')) == 2 :
                    next_page = soup.select('a.pn')[1].get('href')
                elif len(soup.select('td > a.pn')) == 1 :
                    if'다음' in soup.select('td > a.pn')[0].text :
                        next_page = soup.select('a.pn')[0].get('href')
                    else: 
                        break
                elif len(soup.select('td > a.pn')) == 0 :
                    break
                    #다음페이지로 이동하는 코드 
                Current_URL = Current_URL+next_page
                    
                    #다음페이지로 넘길 때, 크롤링카운트를 10증가시킴
                crowling_count = crowling_count + 10 
                    
                #데이터의 저장
            wb.save('C:/Users/'+folder_name+'/'+Keyword_list[i]+'.xlsx')
            time.sleep(60)

    #명령문 입력
    news_crowler(Keyword_list1)    
    
              
    #### 뉴스 업로더 ####
    import openpyxl
    from openpyxl import Workbook
    from selenium.webdriver.support.ui import Select

    #로그인되었으면 다음을 주석처리#
    def news_uploader(Keyword_list) : 
        driver.get('http://127.0.0.1:8080/admin/MorningBriefing/news/')
        html = driver.page_source    
        soup = BeautifulSoup(html,'html.parser')
        if soup.select('h1 > a')[0].text == 'Django administration':
            driver.implicitly_wait(100)
            Username = driver.find_element_by_id('id_username')
            Password = driver.find_element_by_name('password')
            Username.send_keys('yungunmon')
            Password.send_keys('dbsrjsahs23')
            Password.submit()
            
        i = 0 ## 다음 키워드로 넘겨주는 변수
        for i in range(len(Keyword_list)): #리스트의 키워드를 반복하기 위한 For 반복문#
            #키워드의 엑셀파일을 열고, 데이터를 불러옴
            wb = openpyxl.load_workbook('C:/Users/'+folder_name+'/'+Keyword_list[i]+'.xlsx')
            Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
            j = 0
            if Today_Sheet.max_column == 4:
                #엑셀의 데이터를 장고로 옮겨주는 알고리즘
                for j in range(Today_Sheet.max_row):
                    driver.get('http://127.0.0.1:8080/admin/MorningBriefing/news/add/')
                    html = driver.page_source    
                    soup = BeautifulSoup(html,'html.parser')
                    driver.implicitly_wait(100)
                
                # 제목/ 링크/ 업로드 버튼을 따옴#
                    title = driver.find_element_by_id('id_Title')
                    link = driver.find_element_by_id('id_Link')
                    newscompany = driver.find_element_by_id('id_Company')
                    newsdate = driver.find_element_by_name('Published_date')
                    data_field = Select(driver.find_element_by_id('id_Data_field'))
                    write =driver.find_element_by_name('_addanother')
                    driver.implicitly_wait(100)
                    title.send_keys(Today_Sheet.cell(row=j+1, column=1).value)
                    driver.implicitly_wait(100)
                    link.send_keys(Today_Sheet.cell(row=j+1, column=2).value)
                    driver.implicitly_wait(100)
                    newscompany.send_keys(Today_Sheet.cell(row=j+1, column=3).value)
                    driver.implicitly_wait(100)
                    newsdate.send_keys(Today_Sheet.cell(row=j+1, column=4).value)
                    
                    if Keyword_list[i] in ['남동발전', '남부발전','서부발전', '중부발전','동서발전','한국전력','전력거래소','한국수력원자력'] :
                        data_field.select_by_value('A')
                    if Keyword_list[i] in ['전력산업','발전소','석탄','복합발전','LNG','수력','SMP','원전','REC'] :
                        data_field.select_by_value('B')
                    if Keyword_list[i] in ['신재생','태양광','풍력','연료전지','전기차','에너지저장장치' ,'우드펠릿','온실가스','미세먼지', '수소','바이오매스', '에너지']:
                        data_field.select_by_value('C')
                    if Keyword_list[i] in ['국제유가','환율','골드만삭스','기재부','산업부','환경부']:
                        data_field.select_by_value('D')
             
                    write.submit()
                    driver.implicitly_wait(100)
            wb.close()
    #명령문입력#
    news_uploader(Keyword_list1)
    

    def keei_crowler():
        url = 'http://m.keei.re.kr/mobile.nsf/index?readform'
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html,'html.parser')
        driver.implicitly_wait(100)
        i=0
        titles = []
        wb = openpyxl.load_workbook('C:/Users/'+folder_name+'/'+'에너지경제연구원'+'.xlsx')
        wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
        Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
        for i in range(4):
            titles = soup.select("td.ellipsis ")[2*i].text
            date = soup.select("td.ellipsis")[2*i+1].text
            dates=date[4:14]
            link= 'http://www.keei.re.kr/main.nsf/index_mobile.html'
            Today_Sheet.cell(row=i + 1,column=1).value=titles
            Today_Sheet.cell(row=i + 1,column=2).value=dates
            Today_Sheet.cell(row=i + 1,column=3).value=link
            i+=1
        wb.save('C:/Users/'+folder_name+'/'+'에너지경제연구원'+'.xlsx')

    def LG_crowler():
        url = 'http://www.lgeri.com/latest/list.do'
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html,'html.parser')
        driver.implicitly_wait(100)
        wb = openpyxl.load_workbook('C:/Users/'+folder_name+'/'+'LG경제연구원'+'.xlsx')
        wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
        Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
        for i in range(10):
            titles = soup.select(".tit.dotdotdot")[i].text
            titles = titles.strip()
            date = soup.select(".info>span:nth-child(2)")[i].text
            date = date.replace('.','-')
            link ="http://www.lgeri.com/report/view.do?idx="+ soup.select("input[type='hidden']:nth-child(1)")[i+1].get('value')
            Today_Sheet.cell(row=i + 1,column=1).value=titles
            Today_Sheet.cell(row=i + 1,column=2).value=date
            Today_Sheet.cell(row=i + 1,column=3).value=link
            i+=1
        wb.save('C:/Users/'+folder_name+'/'+'LG경제연구원'+'.xlsx')
    def kotra_crowler():
        wb = openpyxl.load_workbook('C:/Users/' + folder_name + '/' + '코트라' + '.xlsx')
        wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
        Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
        for j in range(2):
            url = [
                'http://news.kotra.or.kr/user/globalAllBbs/kotranews/list/2/globalBbsDataAllList.do?column=&search=&searchCmmdyCateCd=&searchAreaCd=&searchNationCd=&searchTradeCd=&searchStartDate=&searchEndDate=&searchIndustryCateIdx=&row=40&CSRFToken=64abbdcf-2f07-4033-9e1b-3a733ecf63ad',
                'http://news.kotra.or.kr/user/globalAllBbs/kotranews/list/781/globalBbsDataAllList.do?column=&search=&searchCmmdyCateCd=&searchAreaCd=&searchNationCd=&searchTradeCd=&searchStartDate=&searchEndDate=&searchIndustryCateIdx=&row=40&CSRFToken=64abbdcf-2f07-4033-9e1b-3a733ecf63ad']
            driver.get(url[j])
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            driver.implicitly_wait(100)
            
            for i in range(20):
                title = soup.select("tbody > tr >td.al")[i].text
                titles = title.strip()
                dates = soup.select("tr>td:nth-child(7)")[i].text
                date = dates.strip()
                link = 'http://news.kotra.or.kr/' + soup.select("tbody > tr >td.al> a")[i].get('href')
                Today_Sheet.cell(row=j * 20 + i + 1, column=1).value = titles
                Today_Sheet.cell(row=j * 20 + i + 1, column=2).value = date
                Today_Sheet.cell(row=j * 20 + i + 1, column=3).value = link
                i += 1
            wb.save('C:/Users/' + folder_name + '/' + '코트라' + '.xlsx')
            
        for k in range(2):
            url = [
                'http://news.kotra.or.kr/user/reports/kotranews/20/usrReportsList.do?page=1&reportsIdx=&hotClipType=&orderByType=list&searchStartDate=&searchEndDate=&searchReportGbn=title&searchText=&searchAreaCd=&searchIndustryCateIdx=&CSRFToken=9898f96f-5f2f-4795-b6ac-f15f3ed24492',
                'http://news.kotra.or.kr/user/reports/kotranews/679/usrReportsList.do?page=1&reportsIdx=&hotClipType=TRADE&orderByType=list&searchStartDate=&searchEndDate=&searchReportGbn=title&searchText=&searchAreaCd=&searchIndustryCateIdx=&CSRFToken=9898f96f-5f2f-4795-b6ac-f15f3ed24492'
            ]
            driver.get(url[k])
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            driver.implicitly_wait(100)
            for l in range(12):
                title = soup.select("tbody > tr > td.al")[l].text
                titles = title.strip()
                dates = soup.select("tbody > tr > td:nth-child(3)")[l].text
                date = dates.strip()
                link = soup.select("tbody > tr > td.al > a")[l].get('href')
                link = link[22:27]
                links = 'http://news.kotra.or.kr/user/reports/kotranews/679/usrReportsView.do?page=1&reportsIdx=' + link + '&hotClipType=TRADE&orderByType=list&searchStartDate=&searchEndDate=&searchReportGbn=title&searchText=&searchAreaCd=&searchIndustryCateIdx=&CSRFToken=9898f96f-5f2f-4795-b6ac-f15f3ed24492'
                Today_Sheet.cell(row=40 + k * 12 + l + 1, column=1).value = titles
                Today_Sheet.cell(row=40 + k * 12 + l + 1, column=2).value = date
                Today_Sheet.cell(row=40 + k * 12 + l + 1, column=3).value = links
                l +=1
            wb.save('C:/Users/' + folder_name + '/' + '코트라' + '.xlsx')
    def HRI_crowler():
        url = 'http://www.hri.co.kr/storage/newReList.asp'
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html,'html.parser')
        driver.implicitly_wait(100)
        wb = openpyxl.load_workbook('C:/Users/'+folder_name+'/'+'현대경제연구원'+'.xlsx')
        wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
        Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
        for i in range(15):
            titles = soup.select("td.textLeft >a:nth-child(2)")[i].text
            date = soup.select("td:nth-child(3)")[i].text
            date = date.replace('.','-')
            link = 'http://www.hri.co.kr/storage/newReList.asp'
            Today_Sheet.cell(row=i + 1,column=1).value=titles
            Today_Sheet.cell(row=i + 1,column=2).value=date
            Today_Sheet.cell(row=i + 1,column=3).value=link
            i+=1
        wb.save('C:/Users/'+folder_name+'/'+'현대경제연구원'+'.xlsx')


    def posco_crowler():
        url = 'https://www.posri.re.kr/ko/board/detailsearch'
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        driver.implicitly_wait(100)
        wb = openpyxl.load_workbook('C:/Users/'+folder_name+'/'+'포스코경영연구원'+'.xlsx')
        wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
        Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
       
        for i in range(len(soup.select("h4 > a"))):
            titles = soup.select("h4 > a")[i].text
            titles = titles.strip()
            date = soup.select('.detail_date')[i].text
            date = date[3:]
            date = date.replace('.','-')
            link_partition= soup.select('h4 > a')[i].get('href')
            link = 'https://www.posri.re.kr/'+link_partition
            Today_Sheet.cell(row=i + 1,column=1).value=titles
            Today_Sheet.cell(row=i + 1,column=2).value=date
            Today_Sheet.cell(row=i + 1,column=3).value=link
            i+=1
        wb.save('C:/Users/'+folder_name+'/'+'포스코경영연구원'+'.xlsx')


    def kepco_crowler():
        wb = openpyxl.load_workbook('C:/Users/'+folder_name+'/'+'한전경제연구원'+'.xlsx')
        wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
        Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
        for j in range(6):
            url = ['https://home.kepco.co.kr/kepco/KR/ntcob/list.do?boardCd=BRD_000271&menuCd=FN310201'
                  ,'https://home.kepco.co.kr/kepco/KR/ntcob/list.do?boardCd=BRD_000353&menuCd=FN310202'
                  ,'https://home.kepco.co.kr/kepco/KR/ntcob/list.do?boardCd=BRD_000354&menuCd=FN310203'
                  ,'https://home.kepco.co.kr/kepco/KR/ntcob/list.do?boardCd=BRD_000355&menuCd=FN310204'
                  ,'https://home.kepco.co.kr/kepco/KR/ntcob/list.do?boardCd=BRD_000356&menuCd=FN310205'
                  ,'https://home.kepco.co.kr/kepco/KR/ntcob/list.do?boardCd=BRD_000357&menuCd=FN310206'
                  ]
            driver.get(url[j])
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            driver.implicitly_wait(100)
            
            for i in range(10):
                titles = soup.select(".tit")[i].text
                titles = titles.strip()
                date = soup.select("tr > td:nth-child(4)")[i].text
                date = date.replace('.','-')
                link_partition_1 = 'https://home.kepco.co.kr/kepco/KR/ntcob/ntcobView.do?pageIndex=1&boardSeq='
                link_partition_2 = soup.select(" td.tit > a")[i].get('onclick')
                link_partition_3 = '&boardCd=BRD_000271&menuCd=FN310201&parnScrpSeq=0&searchCondition=total&searchKeyword='
                link = link_partition_1 +link_partition_2[11:19] +link_partition_3
                Today_Sheet.cell(row=j*10 + i + 1,column=1).value=titles
                Today_Sheet.cell(row=j*10 + i + 1,column=2).value=date
                Today_Sheet.cell(row=j*10 + i + 1,column=3).value=link
                i+=1
            wb.save('C:/Users/'+folder_name+'/'+'한전경제연구원'+'.xlsx')

    def keri_crowler():
        url = 'http://keri.koreaexim.go.kr/site/main/index007'
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        driver.implicitly_wait(100)
        wb = openpyxl.load_workbook('C:/Users/'+folder_name+'/'+'수출입은행해외경제연구소'+'.xlsx')
        wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
        Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
        for i in range(4):
            titles = soup.select("div.new_r > ul > li >a")[i].text
            titles = titles.strip()
            dates = soup.select("div.new_r > ul > li ")[i].text
            datet = dates[50:100]
            date = datet.strip()
            date = date.replace('.','-')
            link_partition = soup.select(" div.new_r > ul > li > a")[i].get('href')
            link = 'http://keri.koreaexim.go.kr' + link_partition
            Today_Sheet.cell(row=i + 1,column=1).value=titles
            Today_Sheet.cell(row=i + 1,column=2).value=date
            Today_Sheet.cell(row=i + 1,column=3).value=link
            i+=1
            wb.save('C:/Users/'+folder_name+'/'+'수출입은행해외경제연구소'+'.xlsx')
            
        for j in range(4):
            titles1 = soup.select("div.issue > ul > li > a")[j].text
            titles1 = titles.strip()
            dates1 = soup.select("div.issue > ul > li")[j].text
            datet1 = dates1[50:100]
            date1 = datet1.strip()
            date1 = date1.replace('.','-')
            link_partition1 = soup.select("div.issue > ul > li > a")[j].get('href')
            link1 = 'http://keri.koreaexim.go.kr' + link_partition1
            Today_Sheet.cell(row=j + 5, column=1).value = titles1
            Today_Sheet.cell(row=j + 5, column=2).value = date1
            Today_Sheet.cell(row=j + 5, column=3).value = link1
            j +=1

        wb.save('C:/Users/'+folder_name+'/'+'수출입은행해외경제연구소'+'.xlsx')

    def samsung_crowler():
        wb = openpyxl.load_workbook('C:/Users/' + folder_name + '/' + '삼성경제연구소' + '.xlsx')
        wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
        Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
        for j in range(2):
            url =[
                'http://www.seri.org/ic/icRPdsZoneL.html?g_menu=06&s_menu=0614',
                'http://www.seri.org/ic/icRPdsZoneL.html?g_menu=06&s_menu=0613'
            ]
            driver.get(url[j])
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            driver.implicitly_wait(100)

            for i in range(10):
                titles = soup.select("td.tit > a.txtlist")[i].text
                titles = titles.strip()
                date = soup.select("table.board_list > tbody > tr > td:nth-child(3) ")[i].text
                date = date.replace('.', '-')
                link_partition = soup.select(" td.tit > a.txtlist ")[i].get('href')
                link = 'http://www.seri.org/ic/' + link_partition
                Today_Sheet.cell(row=j * 10 + i + 1, column=1).value = titles
                Today_Sheet.cell(row=j * 10 + i + 1, column=2).value = date
                Today_Sheet.cell(row=j * 10 + i + 1, column=3).value = link
                i += 1
            wb.save('C:/Users/' + folder_name + '/' + '삼성경제연구소' + '.xlsx')
            
    def kdi_crowler():
        url = 'http://www.kdi.re.kr/research/report_all.jsp'
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        driver.implicitly_wait(100)
        wb = openpyxl.load_workbook('C:/Users/'+folder_name+'/'+'한국개발연구원'+'.xlsx')
        wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
        Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
        for i in range(10):
            titles = soup.select("li > div:nth-child(1)> a ")[i].text
            titles = titles.strip()
            dates = soup.select("li > div:nth-child(1)> div.rpt_sup > span ")[2*i].text
            date = dates.replace('/','-')
            links = soup.select("li > div:nth-child(1)> a ")[i].get('href')
            link = 'http://www.kdi.re.kr/research'+ links[1:]
            
            Today_Sheet.cell(row=i + 1,column=1).value=titles
            Today_Sheet.cell(row=i + 1,column=2).value=date
            Today_Sheet.cell(row=i + 1,column=3).value=link
            i+=1
            wb.save('C:/Users/'+folder_name+'/'+'한국개발연구원'+'.xlsx')
            

    def h2news_crowler():
        wb = openpyxl.load_workbook('C:/Users/' + folder_name + '/' + '월간수소경제' + '.xlsx')
        wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
        Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
        for j in range(3):
            url =[
                'http://www.h2news.kr/news/section_list_all.html?sec_no=5',
                'http://www.h2news.kr/news/section_list_all.html?sec_no=7',
                'http://www.h2news.kr/news/section_list_all.html?sec_no=3'
            ]
            driver.get(url[j])
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            driver.implicitly_wait(100)

            for i in range(10):
                titles = soup.select("div > ul.art_list_all > li > a > h2")[i].text
                titles = titles.strip()
                date = soup.select("ul.art_list_all > li > a > ul > li:nth-child(2) ")[i].text
                date = date.replace('.', '-')
                date = date[0:10]
                link_partition = soup.select(" ul.art_list_all > li > a ")[i].get('href')
                link_partition = link_partition[22:26]
                link = 'http://www.h2news.kr/news/article.html?no=' + link_partition
                Today_Sheet.cell(row=j * 10 + i + 1, column=1).value = titles
                Today_Sheet.cell(row=j * 10 + i + 1, column=2).value = date
                Today_Sheet.cell(row=j * 10 + i + 1, column=3).value = link
                i += 1
            wb.save('C:/Users/' + folder_name + '/' + '월간수소경제' + '.xlsx')
            

    #명령문 입력# 
    #keei_crowler()
    LG_crowler()
    kotra_crowler() 
    HRI_crowler()
    samsung_crowler()
    keri_crowler()
    kepco_crowler()
    posco_crowler()
    kdi_crowler()
    h2news_crowler()

    def report_uploader(report_list) : 
        driver.get('localhost:8080/admin')
        html = driver.page_source    
        soup = BeautifulSoup(html,'html.parser')
        if soup.select('h1 > a')[0].text == 'Django administration':
            driver.implicitly_wait(100)
            Username = driver.find_element_by_id('id_username')
            Password = driver.find_element_by_name('password')
            Username.send_keys('yungunmon')
            Password.send_keys('dbsrjsahs23')
            Password.submit()
        i = 0 ## 다음 키워드로 넘겨주는 변수
        for i in range(len(report_list)): #리스트의 키워드를 반복하기 위한 For 반복문#
            #키워드의 엑셀파일을 열고, 데이터를 불러옴
            wb = openpyxl.load_workbook('C:/Users/'+folder_name+'/'+report_list[i]+'.xlsx')
            Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
            j = 0
            if Today_Sheet.max_column == 3:
                #엑셀의 데이터를 장고로 옮겨주는 알고리즘
                for j in range(Today_Sheet.max_row):
                    driver.get('localhost:8080/admin/MorningBriefing/report/add/')
                    html = driver.page_source    
                    soup = BeautifulSoup(html,'html.parser')
                    driver.implicitly_wait(100)
                    
                    # 제목/ 링크/ 업로드 버튼을 따옴#
                    
                    site = driver.find_element_by_id('id_site')
                    title = driver.find_element_by_id('id_title')
                    link = driver.find_element_by_id('id_link')
                    date = driver.find_element_by_id('id_published_date')
                  
                    site.send_keys(report_list[i])
                    driver.implicitly_wait(100)
                    title.send_keys(Today_Sheet.cell(row=j+1, column=1).value)
                    driver.implicitly_wait(100)
                    link.send_keys(Today_Sheet.cell(row=j+1, column=3).value)
                    driver.implicitly_wait(100)
                    date.send_keys(Today_Sheet.cell(row=j+1, column=2).value) 
                    date.submit()
                    driver.implicitly_wait(100)
            
        wb.close()
        
    report_uploader(report_list)

schedule.every().day.at('02:00').do(job)

while True:
    schedule.run_pending()
    time.sleep(1)


