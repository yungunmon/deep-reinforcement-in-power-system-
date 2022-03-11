from django.shortcuts import render, redirect,get_object_or_404
from .models import News, Report
from django.http import HttpResponse, HttpResponseRedirect,JsonResponse
from .forms import CommentForm, NewTitleForm
from django.views.decorators.csrf import csrf_exempt
import json
from django.template import loader
from datetime import timedelta
import datetime
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from selenium import webdriver
from bs4 import BeautifulSoup
import re
import os
import shutil
import timedelta
import threading
import time
from openpyxl import Workbook
import openpyxl
import schedule


global folder_name

folder_name="svr"

def index(request):
	global start_date, end_date
	today=datetime.date.today()
	one_day_ago = today-datetime.timedelta(days=1)
	try:
		start_date = request.POST["start_date"]
	except:
		start_date = "2019-01-01"
	try:
		end_date=request.POST["end_date"]
	except:
		end_date=today

	
	start_date=start_date[:] if start_date else one_day_ago
	end_date=end_date if end_date else today

	
	news = News.objects.all()
	context = {'news' :news}
	return render(request, 'MorningBriefing/index.html',context)


def Powerinsight(request):
	today=datetime.date.today()
	one_day_ago = today-datetime.timedelta(days=1)
	one_month_ago = today-datetime.timedelta(days=30)
	try:
		start_date=request.POST["start_date"]
	except:
		start_date="2019-01-01"
	try:
		end_date=request.POST["end_date"]
	except:
		end_date=today
	try:
		search=request.POST["search"]
	except:
		search=""
	start_date=start_date if start_date else one_month_ago
	end_date=end_date if end_date else today
	report = Report.objects.filter(Q(published_date__gte=start_date, published_date__lte=end_date)&Q(title__contains=search)).order_by('-published_date')
	context = {'report':report}
	return render(request, 'MorningBriefing/Powerinsight.html',context)


@csrf_exempt
def a(request):
	
	try:
		search=request.POST["search"]
	except:
		search=""


	news_list = News.objects.filter(Q(Data_field='A')&Q(Published_date__gte=start_date, Published_date__lte=end_date)&Q(Title__contains=search)).order_by('-Published_date')
	paginator = Paginator(news_list,20)
	page = request.GET.get('page')

	try:
		news = paginator.page(page)
	except PageNotAnInteger:   
		news = paginator.page(1)
	except EmptyPage:
		news = paginator.page(paginator.num_pages)

	total_len = len(news)
	index = news.number -1 
	max_index = len(paginator.page_range) 
	start_index = index -2 if index >= 2 else 0 
	if index < 2 : 
		end_index = 5-start_index 
	else :
	    end_index = index+3 if index <= max_index - 3 else max_index 
	page_range = list(paginator.page_range[start_index:end_index])


	context = {'news':news, 'page_range':page_range, 'total_len':total_len, 'max_index':max_index-2 ,}
	return render(request, 'MorningBriefing/a.html',context)
	
def b(request):
	
	try:
		search=request.POST["search"]
	except:
		search=""


	news_list = News.objects.filter(Q(Data_field='B')&Q(Published_date__gte=start_date, Published_date__lte=end_date)&Q(Title__contains=search)).order_by('-Published_date')
	paginator = Paginator(news_list,20)
	page = request.GET.get('page')

	try:
		news = paginator.page(page)
	except PageNotAnInteger:   
		news = paginator.page(1)
	except EmptyPage:
		news = paginator.page(paginator.num_pages)

	total_len = len(news)
	index = news.number -1 
	max_index = len(paginator.page_range) 
	start_index = index -2 if index >= 2 else 0 
	if index < 2 : 
		end_index = 5-start_index 
	else :
	    end_index = index+3 if index <= max_index - 3 else max_index 
	page_range = list(paginator.page_range[start_index:end_index])


	context = {'news':news, 'page_range':page_range, 'total_len':total_len, 'max_index':max_index-2 ,}
	return render(request, 'MorningBriefing/b.html',context)

def c(request):
	
	try:
		search=request.POST["search"]
	except:
		search=""


	news_list = News.objects.filter(Q(Data_field='C')&Q(Published_date__gte=start_date, Published_date__lte=end_date)&Q(Title__contains=search)).order_by('-Published_date')
	paginator = Paginator(news_list,20)
	page = request.GET.get('page')

	try:
		news = paginator.page(page)
	except PageNotAnInteger:   
		news = paginator.page(1)
	except EmptyPage:
		news = paginator.page(paginator.num_pages)

	total_len = len(news)
	index = news.number -1 
	max_index = len(paginator.page_range) 
	start_index = index -2 if index >= 2 else 0 
	if index < 2 : 
		end_index = 5-start_index 
	else :
	    end_index = index+3 if index <= max_index - 3 else max_index 
	page_range = list(paginator.page_range[start_index:end_index])


	context = {'news':news, 'page_range':page_range, 'total_len':total_len, 'max_index':max_index-2 ,}
	return render(request, 'MorningBriefing/c.html',context)

def d(request):
	
	try:
		search=request.POST["search"]
	except:
		search=""


	news_list = News.objects.filter(Q(Data_field='D')&Q(Published_date__gte=start_date, Published_date__lte=end_date)&Q(Title__contains=search)).order_by('-Published_date')
	paginator = Paginator(news_list,20)
	page = request.GET.get('page')

	try:
		news = paginator.page(page)
	except PageNotAnInteger:   
		news = paginator.page(1)
	except EmptyPage:
		news = paginator.page(paginator.num_pages)

	total_len = len(news)
	index = news.number -1 
	max_index = len(paginator.page_range) 
	start_index = index -2 if index >= 2 else 0 
	if index < 2 : 
		end_index = 5-start_index 
	else :
	    end_index = index+3 if index <= max_index - 3 else max_index 
	page_range = list(paginator.page_range[start_index:end_index])


	context = {'news':news, 'page_range':page_range, 'total_len':total_len, 'max_index':max_index-2 ,}
	return render(request, 'MorningBriefing/d.html',context)

def report(request):
	news = News.objects.filter(Display= 'A')
	context = {'news':news,}
	return render(request, 'MorningBriefing/report.html',context)

def reportinitialize(request):
	news = News.objects.filter(Display ='A')
	for news in news:
		news.Display = 'B'
		news.save()
	
	return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def comment_detail(request,pk):
	news = get_object_or_404(News,pk=pk)
	return render(request, 'MorningBriefing/comment_detail.html',{'news':news})



def add_comment_to_post(request,pk):
	news = get_object_or_404(News,pk=pk)

	if request.method == 'GET':
		request.session['comment_form'] = request.META.get('HTTP_REFERER', '/')
	if request.method == "POST":
		comment_form = CommentForm(request.POST)


		if CommentForm(request.POST).is_valid():
			news = News.objects.get(pk=pk)
			comment_form = CommentForm(request.POST, instance = news)
			comment_form.save()
			
			if news.Data_field == 'A':
				return HttpResponseRedirect(request.session['comment_form'])

			if news.Data_field == 'B':
				return HttpResponseRedirect(request.session['comment_form'])

			if news.Data_field == 'C':
				return HttpResponseRedirect(request.session['comment_form'])

			if news.Data_field == 'D':
				return HttpResponseRedirect(request.session['comment_form'])
			
	else:
		form = CommentForm()
	return render(request, 'MorningBriefing/add_comment_to_post.html',{'news':news, 'form':form})



def Display_on(request,pk):
	news = get_object_or_404(News,pk=pk)

	if news.Display == 'B':
		news.Display = 'A'
		news.save()

	return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
	

def Display_off(request,pk):
	news = get_object_or_404(News,pk=pk)
	if news.Display == 'A':
		news.Display = 'B'
		news.save()

	return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
def delete(request,pk):
	news = get_object_or_404(News,pk=pk)
	news.Comment = ''
	news.save()

	return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
def delete_news(request,pk):
	
	try:
		search=request.POST["search"]
	except:
		search=""
	news_list = News.objects.filter(Q(Published_date__gte=start_date, Published_date__lte=end_date)&Q(Title__contains=search)).order_by('-Published_date')
	paginator = Paginator(news_list,20)
	page = request.GET.get('page')


	try:
		news = paginator.page(page)
	except PageNotAnInteger:
		news = paginator.page(1)
	except EmptyPage:
		news = paginator.page(paginator.num_pages)
	news = get_object_or_404(News,pk=pk)
	news.delete()
	news.save()

	return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
def modify_title(request,pk):

	try:
		news = paginator.page(page)
	except PageNotAnInteger:
		news = paginator.page(1)
	except EmptyPage:
		news = paginator.page(paginator.num_pages)
	news = get_object_or_404(News,pk=pk)

	if request.method == "POST":
		new_title = NewTitleForm(request.POST)

		if NewTitleForm(request.POST).is_valid():
			news = News.objects.get(pk=pk)
			new_title = NewTitleForm(request.POST, instance = news)
			new_title.save()

			if news.Data_field == 'A':
				return HttpResponseRedirect("/발전사동향/?page="+page)
			if news.Data_field == 'B':
				return HttpResponseRedirect("/전력산업/시장")
			if news.Data_field == 'C':
				return HttpResponseRedirect("/신재생에너지/기술")
			if news.Data_field == 'D':
				return HttpResponseRedirect("/경제/에너지")
	else:
		form = NewTitleForm()
	return render(request, 'MorningBriefing/modify_title.html',{'news':news, 'form':form})


def insert(request):
	news = request.POST.get(pk=pk)
	toSave = models.News(completed=completed)
	toSave.save()
	return render(request,'/')

def crawl(request):
	#======== 셀 레 니 움 설 정 ========#
	options = webdriver.ChromeOptions()
	path = 'C:/Users/svr/chromedriver.exe'
	driver = webdriver.Chrome(path,chrome_options=options)
	#======== 셀 레 니 움 설 정 ========#

	#폴더 생성을 위해 오늘의 날짜를 받아옴
	TODAY_DATE = datetime.datetime.today()
	TTODAY_DATE = TODAY_DATE.strftime("%m/%d/%Y")
	ONEDAY_AGO = (TODAY_DATE - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
	THREE_DAYS_AGO = (TODAY_DATE - datetime.timedelta(days=3)).strftime("%m/%d/%Y")

	#키워드 생성 알고리즘#
	Keyword_list1=['남동발전', '남부발전','서부발전', '중부발전','동서발전','한국전력','전력거래소','한국수력원자력']
	Keyword_list2=['전력산업','발전소','석탄','복합발전','LNG','수력']
	Keyword_list3=['신재생','태양광','풍력','연료전지','전기차','ESS','우드펠릿','온실가스감축','미세먼지']
	Keyword_list4=['국제유가','환율','골드만삭스','기재부','산업부','환경부','공공기관','경영평가']
	Keyword_list5=['에너지','SMP']
	report_list=['에너지경제연구원','포스코경영연구원','LG경제연구원','한전경제연구원','현대경제연구원','수출입은행해외경제연구소','코트라','삼성경제연구소']
	def news_crowler(Keyword_list, Select) : 
		i = 0 ## 다음 키워드로 넘겨주는 변수
		TODAY_DATE = datetime.datetime.today()
		TTODAY_DATE = TODAY_DATE.strftime("%m/%d/%Y")
		ONEDAY_AGO = (TODAY_DATE - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
		THREE_DAYS_AGO = (TODAY_DATE - datetime.timedelta(days=3)).strftime("%m/%d/%Y")
		select = Select
		for i in range(len(Keyword_list)): #리스트의 키워드를 반복하기 위한 For 반복문#
		
			link_partition_1 = 'https://www.google.com/search?q='
			link_partition_2 = Keyword_list[i]
			link_partition_3 = ', -스포츠 , -손흥민, -농구, -전쟁, -야구, -리버풀, -축구, -軍, -핵무기, -공포, -감독&source=lnms&tbm=nws&tbs=qdr:d&sa=X&ved=0ahUKEwi6wYbD8qTjAhULvbwKHWkUCTg4FBD8BQgRKAI&biw=1036&bih=674'
	
			#link_partition_3은 국내 뉴스만 검색하게끔 만듦.
			Current_URL = link_partition_1 + link_partition_2 + link_partition_3
			crowling_count = 0 #크롤링을 한 횟수를 카운트하기 위해 만들어주는 변수 매번 0으로 갱신
			row = 0 #엑셀 파일을 이어붙이기위해 매번 row를 0으로 갱신
			row_index = 0
				
			while (1):
				driver.get(Current_URL)
				html = driver.page_source    
				soup = BeautifulSoup(html,'html.parser')  
				driver.implicitly_wait(300)
					
				#크롤링한 데이터를 엑셀에 옮겨적는 알고리즘
				for row_index in range(len(soup.select("h3 > a"))):
					titles = soup.select("h3 > a")[row_index].text
					links = soup.select("h3 > a")[row_index].get('href')
					news =  soup.select("div.slp > span")[3*row_index].text
					news = news.strip()
					news = news.replace('(풍자)', '')
					news = news.replace('(보도자료)', '')
					news = news.replace('(블로그)', '')
					print(select)
					if News.objects.filter(Link = links):
						print(1)
					else:
						News.objects.create(Title = titles, Link = links, Company = news, Published_date = ONEDAY_AGO, Data_field = select, Comment = ' ')
					#페이지 넘기기 알고리즘
				if len(soup.select('td > a.pn')) == 2 :
					next_page = soup.select('a.pn')[1].get('href')
				elif len(soup.select('td > a.pn')) == 1 :
					if'다음' in soup.select('td > a.pn')[0].text :
						next_page = soup.select('a.pn')[0].get('href')
					else: 
						break
				elif len(soup.select('td > a.pn')) == 0 :
					break
					#다음페이지로 이동하는 코드 
				Current_URL = Current_URL+next_page
					
					#다음페이지로 넘길 때, 크롤링카운트를 10증가시킴
				crowling_count = crowling_count + 10 
	news_crowler(Keyword_list1, 'A')
	time.sleep(300)
	news_crowler(Keyword_list2, 'B')
	time.sleep(300)
	news_crowler(Keyword_list3, 'C')
	time.sleep(300)
	news_crowler(Keyword_list4, 'D')
	time.sleep(300)
	news_crowler(Keyword_list5, 'E')

def power(request):
	from selenium import webdriver
	from bs4 import BeautifulSoup
	from openpyxl import Workbook
	import openpyxl
	import re
	import datetime
	import os
	import shutil

	#======== 셀 레 니 움 설 정 ========#
	options = webdriver.ChromeOptions()
	#헤들리스로 만들기-----------------

	#options.add_argument('headless')
	#options.add_argument('--disable-gpu')
	#options.add_argument('lang=ko_KR')
	#----------------------------------
	path = 'C:/Users/LG/chromedriver.exe'
	driver = webdriver.Chrome(path,chrome_options=options)
	#======== 셀 레 니 움 설 정 ========#


	#폴더 생성을 위해 오늘의 날짜를 받아옴
	TODAY_DATE = datetime.datetime.today()

	#키워드 생성 알고리즘#
	Keyword_list = ['발전사', '한전', '5대발전','전력','태양광발전','신재생에너지','석유','석탄','전기세']
	report_list=['에너지경제연구원','포스코경영연구원','LG경제연구원','한전경제연구원','현대경제연구원','수출입은행해외경제연구소','코트라','삼성경제연구소']




	def keei_crowler():
		url = 'http://m.keei.re.kr/mobile.nsf/index?readform'
		driver.get(url)
		html = driver.page_source
		soup = BeautifulSoup(html,'html.parser')
		driver.implicitly_wait(100)
		i=0
		titles = []
		wb = openpyxl.load_workbook('C:/Users'+folder_name+'/'+'에너지경제연구원'+'.xlsx')
		wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
		Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
		for i in range(4):
			titles = soup.select("td.ellipsis ")[2*i].text
			date = soup.select("td.ellipsis")[2*i+1].text
			dates=date[4:14]
			link= 'http://www.keei.re.kr/main.nsf/index_mobile.html'
			Today_Sheet.cell(row=i + 1,column=1).value=titles
			Today_Sheet.cell(row=i + 1,column=2).value=dates
			Today_Sheet.cell(row=i + 1,column=3).value=link
			i+=1
		wb.save('C:/Users'+folder_name+'/'+'에너지경제연구원'+'.xlsx')
	def LG_crowler():
		url = 'http://www.lgeri.com/latest/list.do'
		driver.get(url)
		html = driver.page_source
		soup = BeautifulSoup(html,'html.parser')
		driver.implicitly_wait(100)
		wb = openpyxl.load_workbook('C:/Users'+folder_name+'/'+'LG경제연구원'+'.xlsx')
		wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
		Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
		for i in range(10):
			titles = soup.select(".tit.dotdotdot")[i].text
			titles = titles.strip()
			date = soup.select(".info>span:nth-child(2)")[i].text
			date = date.replace('.','-')
			link ="http://www.lgeri.com/report/view.do?idx="+ soup.select("input[type='hidden']:nth-child(1)")[i+1].get('value')
			Today_Sheet.cell(row=i + 1,column=1).value=titles
			Today_Sheet.cell(row=i + 1,column=2).value=date
			Today_Sheet.cell(row=i + 1,column=3).value=link
			i+=1
		wb.save('C:/Users'+folder_name+'/'+'LG경제연구원'+'.xlsx')
	def kotra_crowler():
		for j in range(2):
			url = ['http://news.kotra.or.kr/user/globalAllBbs/kotranews/list/2/globalBbsDataAllList.do?column=&search=&searchCmmdyCateCd=&searchAreaCd=&searchNationCd=&searchTradeCd=&searchStartDate=&searchEndDate=&searchIndustryCateIdx=&row=40&CSRFToken=64abbdcf-2f07-4033-9e1b-3a733ecf63ad','http://news.kotra.or.kr/user/globalAllBbs/kotranews/list/781/globalBbsDataAllList.do?column=&search=&searchCmmdyCateCd=&searchAreaCd=&searchNationCd=&searchTradeCd=&searchStartDate=&searchEndDate=&searchIndustryCateIdx=&row=40&CSRFToken=64abbdcf-2f07-4033-9e1b-3a733ecf63ad']
			driver.get(url[j])
			html = driver.page_source
			soup = BeautifulSoup(html,'html.parser')
			driver.implicitly_wait(100)
			wb = openpyxl.load_workbook('C:/Users'+folder_name+'/'+'코트라'+'.xlsx')
			wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
			Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
			for i in range(20):
				title = soup.select("tbody > tr >td.al")[i].text
				titles=title.strip()
				dates = soup.select("tr>td:nth-child(7)")[i].text
				date=dates.strip()
				link = 'http://news.kotra.or.kr/'+soup.select("tbody > tr >td.al> a")[i].get('href')
				Today_Sheet.cell(row=j*20+ i + 1,column=1).value=titles
				Today_Sheet.cell(row=j*20+ i + 1,column=2).value=date
				Today_Sheet.cell(row=j*20+ i + 1,column=3).value=link
				i+=1
			wb.save('C:/Users'+folder_name+'/'+'코트라'+'.xlsx')

	def HRI_crowler():
		url = 'http://www.hri.co.kr/storage/newReList.asp'
		driver.get(url)
		html = driver.page_source
		soup = BeautifulSoup(html,'html.parser')
		driver.implicitly_wait(100)
		wb = openpyxl.load_workbook('C:/Users'+folder_name+'/'+'현대경제연구원'+'.xlsx')
		wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
		Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
		for i in range(15):
			titles = soup.select("td.textLeft >a:nth-child(2)")[i].text
			date = soup.select("td:nth-child(3)")[i].text
			date = date.replace('.','-')
			link = 'http://www.hri.co.kr/storage/newReList.asp'
			Today_Sheet.cell(row=i + 1,column=1).value=titles
			Today_Sheet.cell(row=i + 1,column=2).value=date
			Today_Sheet.cell(row=i + 1,column=3).value=link
			i+=1
		wb.save('C:/Users'+folder_name+'/'+'현대경제연구원'+'.xlsx')


	def posco_crowler():
		url = 'https://www.posri.re.kr/ko/board/detailsearch'
		driver.get(url)
		html = driver.page_source
		soup = BeautifulSoup(html, 'html.parser')
		driver.implicitly_wait(100)

		wb = openpyxl.load_workbook('C:/Users'+folder_name+'/'+'포스코경영연구원'+'.xlsx')
		wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
		Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]

		for i in range(len(soup.select("h4 > a"))):
			titles = soup.select("h4 > a")[i].text
			titles = titles.strip()
			date = soup.select('.detail_date')[i].text
			date = date[3:]
			date = date.replace('.','-')
			link_partition= soup.select('h4 > a')[i].get('href')
			link = 'https://www.posri.re.kr/'+link_partition
			Today_Sheet.cell(row=i + 1,column=1).value=titles
			Today_Sheet.cell(row=i + 1,column=2).value=date
			Today_Sheet.cell(row=i + 1,column=3).value=link
			i+=1
		wb.save('C:/Users'+folder_name+'/'+'포스코경영연구원'+'.xlsx')

	def kepco_crowler():
		url = 'https://home.kepco.co.kr/kepco/KR/ntcob/list.do?boardCd=BRD_000271&menuCd=FN310201'
		driver.get(url)
		html = driver.page_source
		soup = BeautifulSoup(html, 'html.parser')
		driver.implicitly_wait(100)
		wb = openpyxl.load_workbook('C:/Users'+folder_name+'/'+'한전경제연구원'+'.xlsx')
		wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
		Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
		for i in range(10):
			titles = soup.select(".tit")[i].text
			titles = titles.strip()
			date = soup.select("tr > td:nth-child(4)")[i].text
			date = date.replace('.','-')
			link_partition_1 = 'https://home.kepco.co.kr/kepco/KR/ntcob/ntcobView.do?pageIndex=1&boardSeq='
			link_partition_2 = soup.select(" td.tit > a")[i].get('onclick')
			link_partition_3 = '&boardCd=BRD_000271&menuCd=FN310201&parnScrpSeq=0&searchCondition=total&searchKeyword='
			link = link_partition_1 +link_partition_2[11:19] +link_partition_3
			Today_Sheet.cell(row=i + 1,column=1).value=titles
			Today_Sheet.cell(row=i + 1,column=2).value=date
			Today_Sheet.cell(row=i + 1,column=3).value=link
			i+=1
		wb.save('C:/Users'+folder_name+'/'+'한전경제연구원'+'.xlsx')


	def keri_crowler():
		url = 'http://keri.koreaexim.go.kr/site/main/index007'
		driver.get(url)
		html = driver.page_source
		soup = BeautifulSoup(html, 'html.parser')
		driver.implicitly_wait(100)
		wb = openpyxl.load_workbook('C:/Users'+folder_name+'/'+'수출입은행해외경제연구소'+'.xlsx')
		wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
		Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
		for i in range(4):
			titles = soup.select("div.new_r > ul > li >a")[i].text
			titles = titles.strip()
			dates = soup.select("div.new_r > ul > li ")[i].text
			datet = dates[50:100]
			date = datet.strip()
			date = date.replace('.','-')
			link_partition = soup.select(" div.new_r > ul > li > a")[i].get('href')
			link = 'http://keri.koreaexim.go.kr' + link_partition
			Today_Sheet.cell(row=i + 1,column=1).value=titles
			Today_Sheet.cell(row=i + 1,column=2).value=date
			Today_Sheet.cell(row=i + 1,column=3).value=link
			i+=1
		wb.save('C:/Users'+folder_name+'/'+'수출입은행해외경제연구소'+'.xlsx')

	def samsung_crowler():
		url = 'http://www.seri.org/ic/icRPdsZoneL.html?g_menu=06&s_menu=0614'
		driver.get(url)
		html = driver.page_source

		soup = BeautifulSoup(html, 'html.parser')
		driver.implicitly_wait(100)
		wb = openpyxl.load_workbook('C:/Users'+folder_name+'/'+'삼성경제연구소'+'.xlsx')
		wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
		Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
		for i in range(10):
			titles = soup.select("td.tit > a.txtlist")[i].text
			titles = titles.strip()
			date = soup.select("table.board_list > tbody > tr > td:nth-child(3) ")[i].text
			date = date.replace('.','-')
			link_partition = soup.select(" td.tit > a.txtlist ")[i].get('href')
			link = 'http://www.seri.org/ic/' + link_partition
			Today_Sheet.cell(row=i + 1,column=1).value=titles
			Today_Sheet.cell(row=i + 1,column=2).value=date
			Today_Sheet.cell(row=i + 1,column=3).value=link
			i+=1
		wb.save('C:/Users'+folder_name+'/'+'삼성경제연구소'+'.xlsx')
	#명령문 입력# 
	keei_crowler()
	LG_crowler()
	kotra_crowler() 
	HRI_crowler()
	samsung_crowler()
	keri_crowler()
	kepco_crowler()
	posco_crowler()

	def report_uploader(report_list) : 
		driver.get('localhost:8000/admin')
		html = driver.page_source    
		soup = BeautifulSoup(html,'html.parser')
		if soup.select('h1 > a')[0].text == 'Django administration':
			driver.implicitly_wait(100)
			Username = driver.find_element_by_id('id_username')
			Password = driver.find_element_by_name('password')
			Username.send_keys('yungunmon')
			Password.send_keys('dbsrjsahs23')
			Password.submit()
		i = 0 ## 다음 키워드로 넘겨주는 변수
		for i in range(len(report_list)): #리스트의 키워드를 반복하기 위한 For 반복문#
			#키워드의 엑셀파일을 열고, 데이터를 불러옴
			wb = openpyxl.load_workbook('C:/Users'+folder_name+'/'+report_list[i]+'.xlsx')
			Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
			j = 0
			if Today_Sheet.max_column == 3:
				#엑셀의 데이터를 장고로 옮겨주는 알고리즘
				for j in range(Today_Sheet.max_row):
					driver.get('localhost:8000/admin/MorningBriefing/report/add/')
					html = driver.page_source    
					soup = BeautifulSoup(html,'html.parser')
					driver.implicitly_wait(100)
					# 제목/ 링크/ 업로드 버튼을 따옴#
					site = driver.find_element_by_id('id_site')
					title = driver.find_element_by_id('id_title')
					link = driver.find_element_by_id('id_link')
					date = driver.find_element_by_id('id_published_date')
					site.send_keys(report_list[i])
					driver.implicitly_wait(100)
					title.send_keys(Today_Sheet.cell(row=j+1, column=1).value)
					driver.implicitly_wait(100)
					link.send_keys(Today_Sheet.cell(row=j+1, column=3).value)
					driver.implicitly_wait(100)
					date.send_keys(Today_Sheet.cell(row=j+1, column=2).value) 
					date.submit()

					driver.implicitly_wait(100)
			wb.close()
	report_uploader(report_list)

