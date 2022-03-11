from django.shortcuts import render, redirect,get_object_or_404
from .models import News, Report, Comment
from django.http import HttpResponse, HttpResponseRedirect,JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.template import loader
from datetime import timedelta
import datetime
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from selenium import webdriver
from bs4 import BeautifulSoup
import re
import os
import shutil
import timedelta
import threading
import time
from openpyxl import Workbook
import openpyxl
import schedule


global folder_name

folder_name="svr"

def index(request):
    global start_date, end_date
    today=datetime.date.today()
    one_day_ago = today-datetime.timedelta(days=1)
    try:
        start_date = request.POST["start_date"]
    except:
        start_date = "2019-01-01"
    try:
        end_date=request.POST["end_date"]
    except:
        end_date=today

    
    start_date=start_date[:] if start_date else one_day_ago
    end_date=end_date if end_date else today

    
    news = News.objects.all()
    context = {'news' :news}
    return render(request, 'MorningBriefing/index.html',context)


def Powerinsight(request):
    today=datetime.date.today()
    one_day_ago = today-datetime.timedelta(days=1)
    one_month_ago = today-datetime.timedelta(days=30)
    try:
        start_date=request.POST["start_date"]
    except:
        start_date="2019-01-01"
    try:
        end_date=request.POST["end_date"]
    except:
        end_date=today
    try:
        search=request.POST["search"]
    except:
        search=""
    start_date=start_date if start_date else one_month_ago
    end_date=end_date if end_date else today
    report = Report.objects.filter(Q(published_date__gte=start_date, published_date__lte=end_date)&Q(title__contains=search)).order_by('-published_date')
    context = {'report':report}
    return render(request, 'MorningBriefing/Powerinsight.html',context)


@csrf_exempt
def a(request):
    
    try:
        search=request.POST["search"]
    except:
        search=""
    global start_date, end_date
    today=datetime.date.today()
    one_day_ago = today-datetime.timedelta(days=1)

    try:
        start_date = request.POST["start_date"]
    except:
        start_date = '2020-05-01'
    try:
        end_date=request.POST["end_date"]
    except:
        end_date=today

    
    start_date=start_date[:] if start_date else one_day_ago
    end_date=end_date if end_date else today



    news_list = News.objects.filter(Q(Data_field='A')&Q(Published_date__gte=start_date, Published_date__lte=end_date)&Q(Title__contains=search)).order_by('-Published_date','-Importance')

    paginator = Paginator(news_list,100)
    page = request.GET.get('page')

    try:
        news = paginator.page(page)
    except PageNotAnInteger:   
        news = paginator.page(1)
    except EmptyPage:
        news = paginator.page(paginator.num_pages)

    total_len = len(news)
    index = news.number -1 
    max_index = len(paginator.page_range) 
    start_index = index -2 if index >= 2 else 0 
    if index < 2 : 
        end_index = 5-start_index 
    else :
        end_index = index+3 if index <= max_index - 3 else max_index 
    page_range = list(paginator.page_range[start_index:end_index])
    comments = Comment.objects.all()
    context = {'news':news, 'page_range':page_range, 'total_len':total_len, 'max_index':max_index-2,'comments':comments}

    return render(request, 'MorningBriefing/a.html',context)
    
def b(request):
    
    try:
        search=request.POST["search"]
    except:
        search=""
    global start_date, end_date
    today=datetime.date.today()
    one_day_ago = today-datetime.timedelta(days=1)
    
    try:
        start_date = request.POST["start_date"]
    except:
        start_date = '2020-05-01'
    try:
        end_date=request.POST["end_date"]
    except:
        end_date=today

    
    start_date=start_date[:] if start_date else one_day_ago
    end_date=end_date if end_date else today

    news_list = News.objects.filter(Q(Data_field='B')&Q(Published_date__gte=start_date, Published_date__lte=end_date)&Q(Title__contains=search)).order_by('-Published_date','-Importance')
    paginator = Paginator(news_list,100)
    page = request.GET.get('page')

    try:
        news = paginator.page(page)
    except PageNotAnInteger:   
        news = paginator.page(1)
    except EmptyPage:
        news = paginator.page(paginator.num_pages)

    total_len = len(news)
    index = news.number -1 
    max_index = len(paginator.page_range) 
    start_index = index -2 if index >= 2 else 0 
    if index < 2 : 
        end_index = 5-start_index 
    else :
        end_index = index+3 if index <= max_index - 3 else max_index 
    page_range = list(paginator.page_range[start_index:end_index])


    comments = Comment.objects.all()
    context = {'news':news, 'page_range':page_range, 'total_len':total_len, 'max_index':max_index-2,'comments':comments}
    return render(request, 'MorningBriefing/b.html',context)

def c(request):
    
    try:
        search=request.POST["search"]
    except:
        search=""

    global start_date, end_date
    today=datetime.date.today()
    one_day_ago = today-datetime.timedelta(days=1)
    
    try:
        start_date = request.POST["start_date"]
    except:
        start_date = '2020-01-01'
    try:
        end_date=request.POST["end_date"]
    except:
        end_date=today

    
    start_date=start_date[:] if start_date else one_day_ago
    end_date=end_date if end_date else today

    news_list = News.objects.filter(Q(Data_field='C')&Q(Published_date__gte=start_date, Published_date__lte=end_date)&Q(Title__contains=search)).order_by('-Published_date','-Importance')
    paginator = Paginator(news_list,100)
    page = request.GET.get('page')

    try:
        news = paginator.page(page)
    except PageNotAnInteger:   
        news = paginator.page(1)
    except EmptyPage:
        news = paginator.page(paginator.num_pages)

    total_len = len(news)
    index = news.number -1 
    max_index = len(paginator.page_range) 
    start_index = index -2 if index >= 2 else 0 
    if index < 2 : 
        end_index = 5-start_index 
    else :
        end_index = index+3 if index <= max_index - 3 else max_index 
    page_range = list(paginator.page_range[start_index:end_index])


    comments = Comment.objects.all()
    context = {'news':news, 'page_range':page_range, 'total_len':total_len, 'max_index':max_index-2,'comments':comments}
    return render(request, 'MorningBriefing/c.html',context)

def d(request):
    
    try:
        search=request.POST["search"]
    except:
        search=""

    global start_date, end_date
    today=datetime.date.today()
    one_day_ago = today-datetime.timedelta(days=1)
    
    try:
        start_date = request.POST["start_date"]
    except:
        start_date = '2020-05-01'
    try:
        end_date=request.POST["end_date"]
    except:
        end_date=today

    
    start_date=start_date[:] if start_date else one_day_ago
    end_date=end_date if end_date else today

    news_list = News.objects.filter(Q(Data_field='D')&Q(Published_date__gte=start_date, Published_date__lte=end_date)&Q(Title__contains=search)).order_by('-Published_date','-Importance')
    paginator = Paginator(news_list,100)
    page = request.GET.get('page')

    try:
        news = paginator.page(page)
    except PageNotAnInteger:   
        news = paginator.page(1)
    except EmptyPage:
        news = paginator.page(paginator.num_pages)

    total_len = len(news)
    index = news.number -1 
    max_index = len(paginator.page_range) 
    start_index = index -2 if index >= 2 else 0 
    if index < 2 : 
        end_index = 5-start_index 
    else :
        end_index = index+3 if index <= max_index - 3 else max_index 
    page_range = list(paginator.page_range[start_index:end_index])


    comments = Comment.objects.all()
    context = {'news':news, 'page_range':page_range, 'total_len':total_len, 'max_index':max_index-2,'comments':comments}
    return render(request, 'MorningBriefing/d.html',context)

def e(request):

    try:
        search=request.POST["search"]
    except:
        search=""
    global start_date, end_date
    today=datetime.date.today()
    one_day_ago = today-datetime.timedelta(days=1)

    try:
        start_date = request.POST["start_date"]
    except:
        start_date = '2020-05-01'
    try:
        end_date=request.POST["end_date"]
    except:
        end_date=today

    
    start_date=start_date[:] if start_date else one_day_ago
    end_date=end_date if end_date else today



    news_list = News.objects.filter(Q(Data_field='A')&Q(Published_date__gte=start_date, Published_date__lte=end_date)&Q(Title__contains=search)).order_by('-Published_date')

    paginator = Paginator(news_list,100)
    page = request.GET.get('page')

    try:
        news = paginator.page(page)
    except PageNotAnInteger:   
        news = paginator.page(1)
    except EmptyPage:
        news = paginator.page(paginator.num_pages)

    news = News.objects.filter(Display='A').order_by('-Published_date')
    comments = Comment.objects.all()
    context = {'news':news, 'comments':comments}
    return render(request, 'MorningBriefing/e.html',context)

@csrf_exempt
def manage_news(request):
    if request.method == "POST":
        pk = request.POST.get('pk')
        newscategory = request.POST.get('newscategory')
        newstitle = request.POST.get('newstitle')
        newstitle = newstitle.replace('<p>','').replace('</p>','').replace('<br>','')
        print(newscategory,newstitle)
        news = get_object_or_404(News,pk=pk)
        news.Data_field = newscategory
        news.Title = newstitle
        news.save()
        return JsonResponse({'pk':pk, 'newstitle':newstitle})

@csrf_exempt
def manage_comments(request):
    if request.method == "POST":
        pk = request.POST.get('pk')
        comment = get_object_or_404(Comment,pk=pk)
        comment_temp = request.POST.get('comment')
        comment_temp = comment_temp.replace('<p>','').replace('</p>','').replace('<br>','')
        comment.content = comment_temp
        comment.save()
        return JsonResponse({'pk':pk,'comment':comment_temp})
        
def report(request):
    news = News.objects.filter(Display= 'A').order_by('-Published_date')
    today=datetime.date.today()
    comments = Comment.objects.all()
    context = {'news':news,'today':today,'comments':comments}
    return render(request, 'MorningBriefing/report.html',context)

def reportinitialize(request):
    news = News.objects.filter(Display ='A')
    comments = Comment.objects.all()
    for comment in comments:
        comment.delete()
    for news in news:
        news.Display = 'B'
        news.save()
    
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def comment_detail(request,pk):
    news = get_object_or_404(News,pk=pk)
    return render(request, 'MorningBriefing/comment_detail.html',{'news':news})

@csrf_exempt
def add_comment(request):
    if request.method == "POST":
        pk = request.POST.get('pk')
        news = get_object_or_404(News,pk=pk)
        comment_temp = request.POST.get('comment')
        comment = Comment.objects.create(content = comment_temp,news=news)
        comment.save()
        content = ''
        return JsonResponse({'comment' : comment.content,'pk':pk})


def add_comment_to_post(request,pk):
    news = get_object_or_404(News,pk=pk)

    if request.method == 'GET':
        request.session['comment_form'] = request.META.get('HTTP_REFERER', '/')
    if request.method == "POST":
        comment_form = CommentForm(request.POST)


        if CommentForm(request.POST).is_valid():
            news = News.objects.get(pk=pk)
            comment_form = CommentForm(request.POST, instance = news)
            comment_form.save()
            
            if news.Data_field == 'A':
                return HttpResponseRedirect(request.session['comment_form'])

            if news.Data_field == 'B':
                return HttpResponseRedirect(request.session['comment_form'])

            if news.Data_field == 'C':
                return HttpResponseRedirect(request.session['comment_form'])

            if news.Data_field == 'D':
                return HttpResponseRedirect(request.session['comment_form'])
            
    else:
        form = CommentForm()
    return render(request, 'MorningBriefing/add_comment_to_post.html',{'news':news, 'form':form})



def Display_on(request,pk):
    news = get_object_or_404(News,pk=pk)

    if news.Display == 'B':
        news.Display = 'A'
        news.save()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    

def Display_off(request,pk):
    news = get_object_or_404(News,pk=pk)
    if news.Display == 'A':
        news.Display = 'B'
        news.save()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


def delete(request,pk):
    try:
        search=request.POST["search"]
    except:
        search=""
    news_list = News.objects.filter(Q(Published_date__gte=start_date, Published_date__lte=end_date)&Q(Title__contains=search)).order_by('-Published_date')
    paginator = Paginator(news_list,20)
    page = request.GET.get('page')


    try:
        news = paginator.page(page)
    except PageNotAnInteger:
        news = paginator.page(1)
    except EmptyPage:
        news = paginator.page(paginator.num_pages)

    comment = get_object_or_404(Comment,pk=pk)
    comment.delete()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def delete_news(request,pk):
    
    try:
        search=request.POST["search"]
    except:
        search=""
    news_list = News.objects.filter(Q(Published_date__gte=start_date, Published_date__lte=end_date)&Q(Title__contains=search)).order_by('-Published_date')
    paginator = Paginator(news_list,20)
    page = request.GET.get('page')


    try:
        news = paginator.page(page)
    except PageNotAnInteger:
        news = paginator.page(1)
    except EmptyPage:
        news = paginator.page(paginator.num_pages)
    news = get_object_or_404(News,pk=pk)
    comments = Comment.objects.filter(news=news)
    for comment in comments:
        comment.delete()
    news.delete()
    
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
def modify_title(request,pk):

    try:
        news = paginator.page(page)
    except PageNotAnInteger:
        news = paginator.page(1)
    except EmptyPage:
        news = paginator.page(paginator.num_pages)
    news = get_object_or_404(News,pk=pk)

    if request.method == "POST":
        new_title = NewTitleForm(request.POST)

        if NewTitleForm(request.POST).is_valid():
            news = News.objects.get(pk=pk)
            new_title = NewTitleForm(request.POST, instance = news)
            new_title.save()

            if news.Data_field == 'A':
                return HttpResponseRedirect("/발전사동향/?page="+page)
            if news.Data_field == 'B':
                return HttpResponseRedirect("/전력산업/시장")
            if news.Data_field == 'C':
                return HttpResponseRedirect("/신재생에너지/기술")
            if news.Data_field == 'D':
                return HttpResponseRedirect("/경제/에너지")
    else:
        form = NewTitleForm()
    return render(request, 'MorningBriefing/modify_title.html',{'news':news, 'form':form})

@csrf_exempt
def news_display(request):
    if request.method =='POST':
        pk = request.POST.get('pk')
        print(pk)
        news = get_object_or_404(News,pk=pk)
        if news.Display == 'B':
            news.Display = 'A'
            news.save()
        else :
            news.Display = 'B'
            news.save() 
        return JsonResponse({'a_id' : "check_"+str(pk), 'display' : news.Display})

def news_expired(request):
    news_all=News.objects.all()
    for news in news_all:
        pub_date = news.Published_date
        day = (pub_date-datetime.date.today()).days
        if day <=-10:
            news.delete()
    return render(request, 'MorningBriefing/index.html')

def insert(request):
    news = request.POST.get(pk=pk)
    toSave = models.News(completed=completed)
    toSave.save()
    return render(request,'/')

def crawl(request):
    #======== 셀 레 니 움 설 정 ========#
    options = webdriver.ChromeOptions()
    path = 'C:/Users/svr/chromedriver.exe'
    driver = webdriver.Chrome(path,chrome_options=options)
    #======== 셀 레 니 움 설 정 ========#

    #폴더 생성을 위해 오늘의 날짜를 받아옴
    TODAY_DATE = datetime.datetime.today()
    TTODAY_DATE = TODAY_DATE.strftime("%m/%d/%Y")
    ONEDAY_AGO = (TODAY_DATE - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
    THREE_DAYS_AGO = (TODAY_DATE - datetime.timedelta(days=3)).strftime("%m/%d/%Y")

    #키워드 생성 알고리즘#
    Keyword_list1=['남동발전', '남동', '남부발전', '남부','서부발전', '서부', '중부발전','중부', '동서발전','동서', '한국전력','전력거래소','한국수력원자력', '발전5사','발전자회사','자회사']
    Keyword_list2=['전력거래소','전력', '전력연구원','전기요금','전기', '에너지 전환', '전력산업','석탄','탈석탄', '탈원전','원전', '석탄화력', '화력', '석탄발전', '발전',
     '화력발전' , '미세먼지발전', '미세먼지' , '온실가스발전', '온실가스', '발전소', '원자력', '전력중개', '송전선로', '송전', '변전소', '계통', '계통연계', '이산화탄소', '발전단가', '원가연동', '원가'
     , '열복합', '계통용량', '핵연료', 'P2G', '폐자원', '스마트발전소', '지역자원시설세', '계통접속', '출력', 'EPC', 'LFG', 'SRF', '가상발전소', '가스터빈', '가스', '가스화력', '개별요금제',
     '기후변화', '기후위기', '디지털변전소', '변전소', '매립가스', '발전보조금', '발전산업', '발전설비', '배출권', '산업부', '석탄재', '소형모듈', 'ESG', '에너지복지', '공급예비율', 'AMI', 'FIP', 'CCUS', 
     '전력생산', '전력설비', '전력수급', '설비', '전력판매', '전류측정', '전류', '제로에너지', '집단에너지', '탄소인증', '탄소중립', '터빈', '파리협정', '폐기물에너지', '폐기물', '저탄장', 'IGCC',
     '탄소세', '에너지믹스', '에너지분권', '에너지소비', '에너지신산업', '에너지자립', '에너지저장', '에너지협력', '에너지효율', '열병합', '예비전력', '원자로', '원전해체', '유연탄', '저탄소', '전기소비',
     '전기연구원', '전기효율', '전력공급', '전력기금', '환경부' ]
    Keyword_list3=['REC', 'SMP','수소경제','수소', '수력','LNG발전','LNG', '신재생','태양광','풍력','연료전지' , '전기차' , 'ESS' , '바이오매스' , '우드펠릿', '지열', '조력', '태양열', '수열', '잉곳',
    'RE100', 'RPS', '그리드패리티', '그린', '그린뉴딜', '그린본드', '그린수소', '녹색', '녹색금융', '드론', '바이오', '바이오혼소', '수소도시', '수소법', '수소산업', '수소차', '청정에너지', '청정연료', 
    '친환경','친환경차', '태양전지', '폐목재', '폴리실리콘', '환경급전', '환경', '환경설비', '바이오차', '소수력', '신에너지', '영농형', '이차전지', '재생에너지', '폐베터리' ]
    Keyword_list4=['WTI','원/달러 환율', '환율', '저유가', '국제유가']
    
    def news_crowler(Keyword_list, Select) : 
        i = 0 ## 다음 키워드로 넘겨주는 변수
        TODAY_DATE = datetime.datetime.today()
        TTODAY_DATE = TODAY_DATE.strftime("%m/%d/%Y")
        ONEDAY_AGO = (TODAY_DATE - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
        THREE_DAYS_AGO = (TODAY_DATE - datetime.timedelta(days=3)).strftime("%m/%d/%Y")
        select = Select
        for i in range(len(Keyword_list)): #리스트의 키워드를 반복하기 위한 For 반복문#
        
            link_partition_1 = 'https://www.google.com/search?q='
            link_partition_2 = Keyword_list[i]
            link_partition_3 = '&lr=lang_ko&tbs=lr:lang_1ko,qdr:d&tbm=nws&source=lnt&sa=X&ved=0ahUKEwiP9r2f7N3qAhVDPnAKHVhdBzgQpwUIIQ&biw=1080&bih=608&dpr=1.25'
            #link_partition_3은 국내 뉴스만 검색하게끔 만듦.link_partition_3은 국내 뉴스만 검색하게끔 만듦.
            Current_URL = link_partition_1 + link_partition_2 + link_partition_3
            crowling_count = 0 #크롤링을 한 횟수를 카운트하기 위해 만들어주는 변수 매번 0으로 갱신
            row = 0 #엑셀 파일을 이어붙이기위해 매번 row를 0으로 갱신
            row_index = 0
            if i%15==14:
                time.sleep(150)    
            while (1):
                driver.get(Current_URL)
                html = driver.page_source    
                soup = BeautifulSoup(html,'html.parser')  
                driver.implicitly_wait(300)
                
                #크롤링한 데이터를 엑셀에 옮겨적는 알고리즘
                
                if soup.find('div', {'id' : 'rso'})== None:
                    break

                else :
                    news_entity = soup.find('div', {'id' : 'rso'}).select('div.dbsr')
                    for j in range(len(news_entity)):           
                        title = news_entity[j].select("div.JheGif")[0].text
                        news = news_entity[j].select("div.XTjFC")[0].text
                        links = news_entity[j].select("a")[0].get('href')
                        if "연합뉴스" in news:
                            importance = 7
                        elif "뉴시스" in news:
                            importance = 6.9
                        elif "뉴스1" in news:
                            importance = 6.8
                        elif "이데일리" in news:
                            importance = 6.7
                        elif "에너지데일리" in news:
                            importance = 6.6
                        elif "이투뉴스" in news:
                            importance = 6.5
                        elif "투데이에너지" in news:
                            importance = 6.4
                        elif "더구루" in news:
                            importance = 6.3
                        elif "전기신문" in news:
                            importance = 6.2
                        elif "전자신문" in news:
                            importance = 6.1
                        elif "에너지경제" in news:
                            importance = 6.0
                        elif "CEO스코어데일리" in news:
                            importance = 5.9
                        elif "EPJ" in news:
                            importance = 5.8
                        elif "아주경제" in news:
                            importance = 5.7
                        elif "ZDnet korea" in news:
                            importance = 5.6
                        elif "매일경제" in news:
                            importance = 5.5
                        elif "녹색경제신문" in news:
                            importance = 5.4
                        elif "뉴스핌" in news:
                            importance = 5.3
                        elif "한겨레" in news:
                            importance = 5.2
                        elif "파이낸셜뉴스" in news:
                            importance = 5.1
                        elif "데일리안" in news:
                            importance = 5.0
                        elif "동아일보" in news:
                            importance = 4.9
                        elif "일렉트릭파워" in news:
                            importance = 4.8
                        elif "서울경제" in news:
                            importance = 4.7
                        elif "EBN" in news:
                            importance = 4.6
                        elif "세계일보" in news:
                            importance = 4.5
                        elif "지앤이타임즈" in news:
                            importance = 4.4
                        elif "데일리그리드" in news:
                            importance = 4.3
                        elif "한국일보" in news:
                            importance = 4.2
                        elif "국민일보" in news:
                            importance = 4.1
                        elif "아시아경제" in news:
                            importance = 4.0
                        elif "서울신문" in news:
                            importance = 3.9
                        elif "가스신문" in news:
                            importance = 3.8
                        elif "노컷뉴스" in news:
                            importance = 3.7
                        elif "헤럴드경제" in news:
                            importance = 3.6
                        elif "그린포스트코리아" in news:
                            importance = 3.5
                        elif "건설경제신문" in news:
                            importance = 3.4
                        elif "비즈니스포스트" in news:
                            importance = 3.3
                        elif "경향신문" in news:
                            importance = 3.2
                        elif "이투데이" in news:
                            importance = 3.1
                        elif "중앙일보" in news:
                            importance = 3.0
                        elif "데일리포스트" in news:
                            importance = 2
                        elif "글로벌경제" in news:
                            importance = 2
                        elif "글로벌이코노믹" in news:
                            importance = 2
                        elif "일요서울" in news:
                            importance = 2
                        elif "쿠키뉴스" in news:
                            importance = 2
                        elif "조선일보" in news:
                            importance = 2
                        elif "더벨" in news:
                            importance = 2
                        elif "머니투데이" in news:
                            importance = 2
                        elif "시사뉴스" in news:
                            importance = 2
                        elif "폴리뉴스" in news:
                            importance = 2
                        elif "뉴스웨이" in news:
                            importance = 2
                        elif "MTN" in news:
                            importance = 2
                        elif "국토일보" in news:
                            importance = 2
                        elif "이뉴스투데이" in news:
                            importance = 2
                        elif "브릿지경제" in news:
                            importance = 2
                        elif "데일리한국" in news:
                            importance = 2
                        elif "아시아타임즈" in news:
                            importance = 2
                        elif "프라임경제" in news:
                            importance = 2
                        elif "아이뉴스24" in news:
                            importance = 2
                        elif "디지털타임즈" in news:
                            importance = 2
                        elif "이코노미스트" in news:
                            importance = 2
                        elif "뉴스메이커" in news:
                            importance = 2
                        elif "한국경제" in news:
                            importance = 2
                        elif "스트레이트뉴스" in news:
                            importance = 2
                        elif "인더스트리뉴스" in news:
                            importance = 2
                        elif "뉴스웍스" in news:
                            importance = 2
                        elif "연합인포맥스" in news:
                            importance = 2                                 
                        else:
                            importance = -1 
                        if News.objects.filter(Link = links):
                            print(1)
                       
                        else:
                            News.objects.create(Title = title, Link = links, Company = news, Published_date = ONEDAY_AGO, Data_field = select, Importance= importance)
                            News.objects.filter(Importance= -1).delete()
                          
                        #페이지 넘기기 알고리즘
                    if len(soup.select('td > a.pn')) == 2 :
                        next_page = soup.select('a.pn')[1].get('href')
                    elif len(soup.select('td > a.pn')) == 1 :
                        if'다음' in soup.select('td > a.pn')[0].text :
                            next_page = soup.select('a.pn')[0].get('href')
                        else: 
                            break
                    elif len(soup.select('td > a.pn')) == 0 :
                        break
                        #다음페이지로 이동하는 코드 
                    Current_URL = Current_URL+next_page
                        
                        #다음페이지로 넘길 때, 크롤링카운트를 10증가시킴
                    crowling_count = crowling_count + 10 
                
    news_crowler(Keyword_list1, 'A')
    time.sleep(150)
    news_crowler(Keyword_list2, 'B')
    time.sleep(150)
    news_crowler(Keyword_list3, 'C')
    time.sleep(150)
    news_crowler(Keyword_list4, 'D')
    return render(request, 'MorningBriefing/index.html')

def power(request):
    from selenium import webdriver
    from bs4 import BeautifulSoup
    
    import re
    import datetime
    import os
    import shutil

    #======== 셀 레 니 움 설 정 ========#
    options = webdriver.ChromeOptions()
    #헤들리스로 만들기-----------------

    #options.add_argument('headless')
    #options.add_argument('--disable-gpu')
    #options.add_argument('lang=ko_KR')
    #----------------------------------
    path = 'C:/Users/svr/chromedriver.exe'
    driver = webdriver.Chrome(path,chrome_options=options)
    #======== 셀 레 니 움 설 정 ========#


    #폴더 생성을 위해 오늘의 날짜를 받아옴
    TODAY_DATE = datetime.datetime.today()

    #키워드 생성 알고리즘#
    Keyword_list = ['발전사', '한전', '5대발전','전력','태양광발전','신재생에너지','석유','석탄','전기세']
    report_list=['에너지경제연구원','포스코경영연구원','LG경제연구원','한전경제연구원','현대경제연구원','수출입은행해외경제연구소','코트라','삼성경제연구소']

    def keei_crowler():
        url = 'http://m.keei.re.kr/mobile.nsf/index?readform'
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html,'html.parser')
        driver.implicitly_wait(100)
        i=0
        titles = []
        wb = openpyxl.load_workbook('C:/Users'+folder_name+'/'+'에너지경제연구원'+'.xlsx')
        wb.create_sheet(TODAY_DATE.strftime('%Y-%m-%d'))
        Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
        for i in range(4):
            titles = soup.select("td.ellipsis ")[2*i].text
            date = soup.select("td.ellipsis")[2*i+1].text
            dates=date[4:14]
            link= 'http://www.keei.re.kr/main.nsf/index_mobile.html'
            
    def LG_crowler():
        url = 'http://www.lgeri.com/latest/list.do'
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html,'html.parser')
        driver.implicitly_wait(100)

        for i in range(10):
            titles = soup.select(".tit.dotdotdot")[i].text
            titles = titles.strip()
            date = soup.select(".info>span:nth-child(2)")[i].text
            date = date.replace('.','-')
            links ="http://www.lgeri.com/report/view.do?idx="+ soup.select("input[type='hidden']:nth-child(1)")[i+1].get('value')
            if Report.objects.filter(title = titles):
                print(1)
            else:
                Report.objects.create(site = 'LG경제연구원' , title = titles, link = links, published_date = date)
            i+=1
        
    def kotra_crowler():
        for j in range(2):
            url = ['http://news.kotra.or.kr/user/globalAllBbs/kotranews/list/2/globalBbsDataAllList.do?column=&search=&searchCmmdyCateCd=&searchAreaCd=&searchNationCd=&searchTradeCd=&searchStartDate=&searchEndDate=&searchIndustryCateIdx=&row=40&CSRFToken=64abbdcf-2f07-4033-9e1b-3a733ecf63ad','http://news.kotra.or.kr/user/globalAllBbs/kotranews/list/781/globalBbsDataAllList.do?column=&search=&searchCmmdyCateCd=&searchAreaCd=&searchNationCd=&searchTradeCd=&searchStartDate=&searchEndDate=&searchIndustryCateIdx=&row=40&CSRFToken=64abbdcf-2f07-4033-9e1b-3a733ecf63ad']
            driver.get(url[j])
            html = driver.page_source
            soup = BeautifulSoup(html,'html.parser')
            driver.implicitly_wait(100)
            
            for i in range(20):
                title = soup.select("tbody > tr >td.al")[i].text
                titles=title.strip()
                dates = soup.select("tr>td:nth-child(7)")[i].text
                date=dates.strip()
                link = 'http://news.kotra.or.kr/'+soup.select("tbody > tr >td.al> a")[i].get('href')
                if Report.objects.filter(title = titles):
                    print(1)
                else:
                    Report.objects.create(site = '코트라' , title = titles, link = link, published_date = date)
                i+=1
                
    def HRI_crowler():
        url = 'http://www.hri.co.kr/storage/newReList.asp'
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html,'html.parser')
        driver.implicitly_wait(100)
        
        for i in range(15):
            titles = soup.select("td.textLeft >a:nth-child(2)")[i].text
            date = soup.select("td:nth-child(3)")[i].text
            date = date.replace('.','-')
            link = 'http://www.hri.co.kr/storage/newReList.asp'
            if Report.objects.filter(title = titles):
                print(1)
            else:
                Report.objects.create(site = '현대경제연구원' , title = titles, link = link, published_date = date)
            i+=1
        

    def posco_crowler():
        url = 'https://www.posri.re.kr/ko/board/detailsearch'
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        driver.implicitly_wait(100)

        for i in range(len(soup.select("h4 > a"))):
            titles = soup.select("h4 > a")[i].text
            titles = titles.strip()
            date = soup.select('.detail_date')[i].text
            date = date[3:]
            date = date.replace('.','-')
            date = date.strip()
            link_partition= soup.select('h4 > a')[i].get('href')
            link = 'https://www.posri.re.kr/'+link_partition
            if Report.objects.filter(title = titles):
                print(1)
            else:
                Report.objects.create(site = '포스코경영연구원' , title = titles, link = link, published_date = date)
            i+=1

    def kepco_crowler():
        url = 'https://home.kepco.co.kr/kepco/KR/ntcob/list.do?boardCd=BRD_000271&menuCd=FN310201'
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        driver.implicitly_wait(100)
        for i in range(10):
            titles = soup.select(".tit")[i].text
            titles = titles.strip()
            date = soup.select("tr > td:nth-child(4)")[i].text
            date = date.replace('.','-')
            link_partition_1 = 'https://home.kepco.co.kr/kepco/KR/ntcob/ntcobView.do?pageIndex=1&boardSeq='
            link_partition_2 = soup.select(" td.tit > a")[i].get('onclick')
            link_partition_3 = '&boardCd=BRD_000271&menuCd=FN310201&parnScrpSeq=0&searchCondition=total&searchKeyword='
            link = link_partition_1 +link_partition_2[11:19] +link_partition_3
            if Report.objects.filter(title = titles):
                print(1)
            else:
                Report.objects.create(site = '한전경제연구원' , title = titles, link = link, published_date = date)
            i+=1


    def keri_crowler():
        url = 'http://keri.koreaexim.go.kr/site/program/board/basicboard/list?boardtypeid=168&menuid=007002001003&boardtext8=PA03'
        driver.get(url)
        html = driver.page_source

        soup = BeautifulSoup(html, 'html.parser')
        driver.implicitly_wait(100)
        news_entity = soup.find('tbody').select('tr')
        print(len(news_entity))

        for i in range(len(news_entity)):
            titles = news_entity[i].select("a")[0].text   
            
            date = news_entity[i].select("td:nth-child(4)")[0].text
            date = date.replace('.','-')
            link_partition = news_entity[i].select(" a ")[0].get('href') 
            link_partition = link_partition.replace('.','')
            link = 'http://keri.koreaexim.go.kr/site/program/board/basicboard' + link_partition
            if Report.objects.filter(title = titles):
                print(1)
            else:
                Report.objects.create(site = '수출입은행해외경제연구소' , title = titles, link = link, published_date = date)
            i+=1

    def samsung_crowler():
        url = 'http://www.seri.org/ic/icRPdsZoneL.html?g_menu=06&s_menu=0614'
        driver.get(url)
        html = driver.page_source

        soup = BeautifulSoup(html, 'html.parser')
        driver.implicitly_wait(100)
        
        for i in range(10):
            titles = soup.select("td.tit > a.txtlist")[i].text
            titles = titles.strip()
            date = soup.select("table.board_list > tbody > tr > td:nth-child(3) ")[i].text
            date = date.replace('.','-')
            link_partition = soup.select(" td.tit > a.txtlist ")[i].get('href')
            link = 'http://www.seri.org/ic/' + link_partition
            if Report.objects.filter(title = titles):
                print(1)
            else:
                Report.objects.create(site = '삼성경제연구소' , title = titles, link = link, published_date = date)
            i+=1


    def ketep_crowler():
        url = 'https://www.ketep.re.kr/brdartcl/boardarticleList.do?brd_id=BDIDX_gn35RzJ24P1R5fU9B0h894&srch_menu_nix=D27pi4kw'
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        driver.implicitly_wait(100)
    
        for i in range(8):
            titles = soup.select("td.ellipsis")[i].text
            titles = titles.strip()
            date = soup.select("tr > td:nth-child(4)")[i].text
            date = date.strip()
            link = 'https://www.ketep.re.kr/brdartcl/boardarticleList.do?brd_id=BDIDX_gn35RzJ24P1R5fU9B0h894&srch_menu_nix=D27pi4kw'
            if Report.objects.filter(title = titles):
                print(1)
            else:
                Report.objects.create(site = '에너지기술평가원' , title = titles, link = link, published_date = date)
            i+=1

    def kdb1_crowler():
        url = 'https://rd.kdb.co.kr/FLPBFP03N01.act?_mnuld=FYERER0031#'
        driver.get(url)
        html = driver.page_source

        soup = BeautifulSoup(html, 'html.parser')
        driver.implicitly_wait(100)

        for i in range(10):
            titles = soup.select("td.al")[i].text
            titles = titles.strip()
            date = soup.select("tr > td:nth-child(4)")[i].text
            date = date.replace('.','-')
            link_partition = soup.select(" td > a ")[i].get('href')
            link = 'https://rd.kdb.co.kr/' + link_partition
            if Report.objects.filter(title = titles):
                print(1)
            else:
                Report.objects.create(site = 'KDB미래전략연구소' , title = titles, link = link, published_date = date)
            i+=1

    def h2news_crowler():
        url = 'https://www.h2news.kr/mobile/section_list_all.html?sec_no=3'
        driver.get(url)
        html = driver.page_source

        soup = BeautifulSoup(html, 'html.parser')
        driver.implicitly_wait(100)
        news_entity = soup.select('li.on_thumb')
        

        for i in range(len(news_entity)):
            titles = news_entity[i].select("h3")[0].text   
            
            date = news_entity[i].select("ul>li:nth-child(1)")[0].text
            date = date[:10]
            link_partition = news_entity[i].select(" a ")[0].get('href') 
            link = 'http://h2news.kr/' + link_partition
            if Report.objects.filter(title = titles):
                print(1)
            else:
                Report.objects.create(site = '월간수소경제' , title = titles, link = link, published_date = date)
            i+=1
    #명령문 입력# 
    #keei_crowler()
    #LG_crowler()
    #kotra_crowler() 
    HRI_crowler()
    #samsung_crowler()
    keri_crowler()
    kepco_crowler()
    posco_crowler()
    ketep_crowler()
    #kdb1_crowler()
    h2news_crowler()
    return render(request, 'MorningBriefing/index.html')

    def report_uploader(report_list) : 
        driver.get('localhost:8000/admin')
        html = driver.page_source    
        soup = BeautifulSoup(html,'html.parser')
        if soup.select('h1 > a')[0].text == 'Django administration':
            driver.implicitly_wait(100)
            Username = driver.find_element_by_id('id_username')
            Password = driver.find_element_by_name('password')
            Username.send_keys('yungunmon')
            Password.send_keys('dbsrjsahs23')
            Password.submit()
        i = 0 ## 다음 키워드로 넘겨주는 변수
        for i in range(len(report_list)): #리스트의 키워드를 반복하기 위한 For 반복문#
            #키워드의 엑셀파일을 열고, 데이터를 불러옴
            wb = openpyxl.load_workbook('C:/Users'+folder_name+'/'+report_list[i]+'.xlsx')
            Today_Sheet = wb[TODAY_DATE.strftime('%Y-%m-%d')]
            j = 0
            if Today_Sheet.max_column == 3:
                #엑셀의 데이터를 장고로 옮겨주는 알고리즘
                for j in range(Today_Sheet.max_row):
                    driver.get('localhost:8000/admin/MorningBriefing/report/add/')
                    html = driver.page_source    
                    soup = BeautifulSoup(html,'html.parser')
                    driver.implicitly_wait(100)
                    # 제목/ 링크/ 업로드 버튼을 따옴#
                    site = driver.find_element_by_id('id_site')
                    title = driver.find_element_by_id('id_title')
                    link = driver.find_element_by_id('id_link')
                    date = driver.find_element_by_id('id_published_date')
                    site.send_keys(report_list[i])
                    driver.implicitly_wait(100)
                    title.send_keys(Today_Sheet.cell(row=j+1, column=1).value)
                    driver.implicitly_wait(100)
                    link.send_keys(Today_Sheet.cell(row=j+1, column=3).value)
                    driver.implicitly_wait(100)
                    date.send_keys(Today_Sheet.cell(row=j+1, column=2).value) 
                    date.submit()

                    driver.implicitly_wait(100)
            wb.close()
#   report_uploader(report_list)

